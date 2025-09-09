# common/xlsx_fill.py
from __future__ import annotations

import io
import os
import re
import unicodedata
from datetime import date, datetime
from typing import Optional, Dict, List, Tuple

from dateutil.relativedelta import relativedelta
import pandas as pd
import openpyxl
from openpyxl.utils import coordinate_to_tuple, get_column_letter

# ======================================================================
#                         Utilitaires généraux
# ======================================================================

VOL_TOL = 0.02

def _is_close(a: float, b: float, tol: float = VOL_TOL) -> bool:
    try:
        return abs(float(a) - float(b)) <= tol
    except Exception:
        return False

# ----------- parse format depuis la colonne "Stock" (df_min) -----------
def _parse_format_from_stock(stock: str):
    s = str(stock or "")
    m_nb = re.search(r'(Carton|Pack)\s+de\s+(\d+)\s+Bouteilles?', s, flags=re.I)
    nb = int(m_nb.group(2)) if m_nb else None
    m_l = re.search(r'(\d+(?:[.,]\d+)?)\s*[lL]\b', s)
    vol = float(m_l.group(1).replace(",", ".")) if m_l else None
    if vol is None:
        m_cl = re.search(r'(\d+(?:[.,]\d+)?)\s*c[lL]\b', s)
        vol = float(m_cl.group(1).replace(",", "."))/100.0 if m_cl else None
    return nb, vol

# ----------- Agrégat STRICT depuis df_min (tableau affiché) -----------
def _agg_from_dfmin(df_min: pd.DataFrame, gout: str) -> Dict[str, Dict[str, int]]:
    out = {
        "33_fr":  {"cartons": 0, "bouteilles": 0},
        "33_niko":{"cartons": 0, "bouteilles": 0},
        "75x6":   {"cartons": 0, "bouteilles": 0},
        "75x4":   {"cartons": 0, "bouteilles": 0},
    }
    if df_min is None or not isinstance(df_min, pd.DataFrame) or df_min.empty:
        return out
    req = {"Produit","Stock","GoutCanon","Cartons à produire (arrondi)","Bouteilles à produire (arrondi)"}
    if any(c not in df_min.columns for c in req):
        return out

    df = df_min.copy()
    df = df[df["GoutCanon"].astype(str).str.strip() == str(gout).strip()]
    if df.empty:
        return out

    for _, r in df.iterrows():
        nb, vol = _parse_format_from_stock(r["Stock"])
        if nb is None or vol is None:
            continue
        ct = int(pd.to_numeric(r["Cartons à produire (arrondi)"], errors="coerce") or 0)
        bt = int(pd.to_numeric(r["Bouteilles à produire (arrondi)"], errors="coerce") or 0)
        prod_up = str(r["Produit"]).upper()

        if nb == 12 and _is_close(vol, 0.33):
            key = "33_niko" if "NIKO" in prod_up else "33_fr"
        elif nb == 6 and _is_close(vol, 0.75):
            key = "75x6"
        elif nb == 4 and _is_close(vol, 0.75):
            key = "75x4"
        else:
            continue

        out[key]["cartons"]    += ct
        out[key]["bouteilles"] += bt

    return out

# ----------- Helper écriture tolérante aux fusions -----------
def _set(ws, addr: str, value, number_format: str | None = None):
    row, col = coordinate_to_tuple(addr)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            row, col = rng.min_row, rng.min_col
            break
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if number_format:
        cell.number_format = number_format
    return f"{get_column_letter(col)}{row}"

# ----------- Détection auto des blocs Quantité (paire du BAS) -----------
def _norm(s) -> str:
    return str(s).strip().lower()

def _locate_quantity_blocks(ws) -> Dict[str, Dict[str, int]]:
    """
    Le modèle contient 2 paires de blocs (haut = résumé, bas = zone d'entrée).
    On retourne **la paire du BAS** pour la saisie.
    """
    labels = {"france", "niko", "x6", "x4"}
    row_hits: Dict[int, Dict[str, int]] = {}

    for r in ws.iter_rows(values_only=False):
        for c in r:
            v = c.value
            if isinstance(v, str):
                nv = _norm(v)
                if nv in labels:
                    row_hits.setdefault(c.row, {})[nv] = c.column

    candidates = [(row, cols) for row, cols in row_hits.items() if len(cols) >= 3]
    if len(candidates) < 2:
        raise KeyError("En-têtes 'France/NIKO/X6/X4' introuvables (paire du bas non détectée).")

    # On prend les 2 lignes les plus basses (bas de page)
    candidates.sort(key=lambda x: x[0])
    bottom_pair = candidates[-2:]

    def _avg_col(cols: Dict[str, int]) -> float:
        return sum(cols.values()) / len(cols)

    # gauche / droite
    bottom_pair.sort(key=lambda x: _avg_col(x[1]))
    (left_row, left_cols), (right_row, right_cols) = bottom_pair

    def _fill_missing(cols: Dict[str, int]) -> Dict[str, int]:
        out = cols.copy()
        for k in ["france", "niko", "x6", "x4"]:
            out.setdefault(k, next(iter(out.values())))
        return out

    left_cols  = _fill_missing(left_cols)
    right_cols = _fill_missing(right_cols)

    return {
        "left":  {"header_row": left_row,  "bouteilles_row": left_row + 1, "cartons_row": left_row + 2, **left_cols},
        "right": {"header_row": right_row, "bouteilles_row": right_row + 1, "cartons_row": right_row + 2, **right_cols},
    }

def _addr(col: int, row: int) -> str:
    return f"{get_column_letter(col)}{row}"

# ======================================================================
#     Fiche de prod 7000L (existante dans ton repo) — inchangée
# ======================================================================

def fill_fiche_7000L_xlsx(
    template_path: str,
    semaine_du: date,
    ddm: date,
    gout1: str,
    gout2: Optional[str],
    df_calc,
    sheet_name: str | None = None,
    df_min=None,
) -> bytes:
    wb = openpyxl.load_workbook(template_path, data_only=False, keep_vba=False)

    targets = [sheet_name] if sheet_name else ["Fiche de production 7000 L", "Fiche de production 7000L"]
    ws = None
    for nm in targets:
        if nm and nm in wb.sheetnames:
            ws = wb[nm]
            break
    if ws is None:
        raise KeyError(f"Feuille cible introuvable. Feuilles présentes : {wb.sheetnames}")

    # En-têtes
    _set(ws, "D8", gout1 or "")
    _set(ws, "T8", gout2 or "")
    _set(ws, "D10", ddm, number_format="DD/MM/YYYY")
    _set(ws, "O10", ddm.strftime("%d%m%Y"))
    ferment_date = ddm - relativedelta(years=1)
    _set(ws, "A20", ferment_date, number_format="DD/MM/YYYY")

    # Localisation des blocs
    blocks = _locate_quantity_blocks(ws)
    L = blocks["left"];  R = blocks["right"]

    P1 = {
        "33_fr":  {"b": _addr(L["france"], L["bouteilles_row"]), "c": _addr(L["france"], L["cartons_row"])},
        "33_niko":{"b": _addr(L["niko"],   L["bouteilles_row"]), "c": _addr(L["niko"],   L["cartons_row"])},
        "75x6":   {"b": _addr(L["x6"],     L["bouteilles_row"]), "c": _addr(L["x6"],     L["cartons_row"])},
        "75x4":   {"b": _addr(L["x4"],     L["bouteilles_row"]), "c": _addr(L["x4"],     L["cartons_row"])},
    }
    P2 = {
        "33_fr":  {"b": _addr(R["france"], R["bouteilles_row"]), "c": _addr(R["france"], R["cartons_row"])},
        "33_niko":{"b": _addr(R["niko"],   R["bouteilles_row"]), "c": _addr(R["niko"],   R["cartons_row"])},
        "75x6":   {"b": _addr(R["x6"],     R["bouteilles_row"]), "c": _addr(R["x6"],     R["cartons_row"])},
        "75x4":   {"b": _addr(R["x4"],     R["bouteilles_row"]), "c": _addr(R["x4"],     R["cartons_row"])},
    }

    # --- Agrégats : df_min uniquement (copie EXACTE du tableau affiché)
    agg1 = _agg_from_dfmin(df_min, gout1)
    agg2 = _agg_from_dfmin(df_min, gout2) if gout2 else None

    # N'écrit rien si 0 → on laisse les pointillés du modèle
    def _write_if_pos(addr: str, val):
        v = int(pd.to_numeric(val, errors="coerce") or 0)
        if v > 0:
            _set(ws, addr, v)

    # Gauche (Produit 1)
    for k, dest in P1.items():
        _write_if_pos(dest["b"], agg1[k]["bouteilles"])
        _write_if_pos(dest["c"], agg1[k]["cartons"])

    # Droite (Produit 2) si présent (sinon on ne touche pas aux pointillés)
    if agg2 is not None:
        for k, dest in P2.items():
            _write_if_pos(dest["b"], agg2[k]["bouteilles"])
            _write_if_pos(dest["c"], agg2[k]["cartons"])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ======================================================================
#                   Remplissage BL enlèvements Sofripa
# ======================================================================

# --- Helpers dédiés au modèle BL Sofripa ---

def _iter_cells(ws):
    for r in ws.iter_rows(values_only=False):
        for c in r:
            yield c

def _find_cell_by_regex(ws, pattern: str) -> Tuple[int, int] | Tuple[None, None]:
    rx = re.compile(pattern, flags=re.I)
    for cell in _iter_cells(ws):
        v = cell.value
        if isinstance(v, str) and rx.search(v):
            return cell.row, cell.column
    return None, None

def _write_right_of(ws, row: int, col: int, value):
    r, c = row, col + 1
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
            r, c = rng.min_row, rng.min_col
            break
    ws.cell(row=r, column=c).value = value


def _normalize_header_text(s: str) -> str:
    s = str(s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace("’", "'")
    for ch in ["(", ")", ":", ";", ","]:
        s = s.replace(ch, " ")
    s = " ".join(s.split())
    return s

def _find_table_headers(ws, targets: List[str]) -> Tuple[int | None, Dict[str, int]]:
    """
    Essaie de trouver une ligne qui ressemble à des en-têtes du tableau principal.
    Retourne (row_index, mapping_nom->col_index_1based)
    """
    norm_targets = [_normalize_header_text(t) for t in targets]

    # on parcourt les premières ~50 lignes pour trouver une majorité de correspondances
    best_row = None
    best_map: Dict[str, int] = {}
    max_hits = 0

    max_rows = min(ws.max_row, 80)
    max_cols = min(ws.max_column, 50)

    for r in range(1, max_rows + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_cols + 1)]
        row_norm = [_normalize_header_text(x) for x in row_vals]

        colmap: Dict[str, int] = {}
        hits = 0
        for t_norm, t_orig in zip(norm_targets, targets):
            found = False
            for j, hv in enumerate(row_norm, start=1):
                if hv == t_norm:
                    colmap[t_orig] = j
                    hits += 1
                    found = True
                    break
            if not found:
                # essais souples (contains)
                for j, hv in enumerate(row_norm, start=1):
                    if t_norm in hv and len(t_norm) >= 4:
                        colmap[t_orig] = j
                        hits += 1
                        found = True
                        break

        if hits > max_hits:
            max_hits = hits
            best_row = r
            best_map = colmap

        if hits >= len(targets) - 1:  # quasi toutes
            break

    return best_row, best_map

# helper: écrit dans la cellule (row,col) en visant l'ancre si c'est une fusion
def _write_cell(ws, row: int, col: int, value):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            row, col = rng.min_row, rng.min_col
            break
    ws.cell(row=row, column=col).value = value

def fill_bl_enlevements_xlsx(
    template_path: str,
    date_creation: date,
    date_ramasse: date,
    destinataire_title: str,
    destinataire_lines: List[str],
    df_lines: pd.DataFrame,   # colonnes attendues (ordre libre) cf. ci-dessous
) -> bytes:
    """
    Remplit le modèle XLSX 'LOG_EN_001_01 BL enlèvements Sofripa-2.xlsx'.

    df_lines doit contenir les colonnes (noms exacts ou équivalents) :
      - 'Référence'
      - 'Produit (goût + format)' ou 'Produit'
      - 'DDM'
      - 'Quantité cartons'
      - 'Quantité palettes'
      - 'Poids palettes (kg)'
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Modèle Excel introuvable: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # ----- 1) Dates -----
    r, c = _find_cell_by_regex(ws, r"date\s+de\s+cr[eé]ation")
    if r and c:
        _write_right_of(ws, r, c, date_creation.strftime("%d/%m/%Y"))

    r, c = _find_cell_by_regex(ws, r"date\s+de\s+rammasse|date\s+de\s+ramasse")
    if r and c:
        _write_right_of(ws, r, c, date_ramasse.strftime("%d/%m/%Y"))

    # ----- 2) Destinataire (robuste : trouve ou crée la fusion à droite) -----
    from openpyxl.styles import Alignment

    r, c = _find_cell_by_regex(ws, r"destinataire")
    if r and c:
        # 1) essaie d'attraper une fusion existante à DROITE du libellé, sur la même bande
        target_rng = None
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= r <= rng.max_row and rng.min_col > c:
                # on privilégie la fusion la plus à gauche (côté valeur)
                if target_rng is None or rng.min_col < target_rng.min_col:
                    target_rng = rng

        if target_rng:
            rr, cc = target_rng.min_row, target_rng.min_col
            rr_end, cc_end = target_rng.max_row, target_rng.max_col
        else:
            # 2) pas de fusion repérée : on EN CRÉE une (3 lignes × 6 colonnes) à droite du libellé
            rr, cc = r, c + 1
            rr_end = min(r + 2, ws.max_row)
            cc_end = min(c + 6, ws.max_column)
            try:
                ws.merge_cells(start_row=rr, start_column=cc, end_row=rr_end, end_column=cc_end)
            except Exception:
                # si la zone est déjà partiellement fusionnée, on ignore l'erreur et on écrira au top-left
                pass

        cell = ws.cell(row=rr, column=cc)

        # 3) écrit TOUT (titre + adresse) en multi-lignes
        lines = [destinataire_title] + (destinataire_lines or [])
        text = "\n".join([str(x).strip() for x in lines if str(x).strip()])
        cell.value = text
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 4) ajuste la hauteur des lignes de la zone pour rendre le texte visible
        n_lines = max(1, text.count("\n") + 1)
        span_rows = max(1, (rr_end - rr + 1))
        # ~14 points par ligne répartis sur les lignes fusionnées
        per_row_height = 14 * n_lines / span_rows
        for rset in range(rr, rr_end + 1):
            cur = ws.row_dimensions[rset].height or 0
            ws.row_dimensions[rset].height = max(cur, per_row_height)

        # 5) (optionnel) si une ligne parasite “ZAC du Haut…” traîne ailleurs, on la vide
        zr, zc = _find_cell_by_regex(ws, r"zac\s+du\s+haut\s+de\s+wissous")
        if zr and zc:
            ws.cell(row=zr, column=zc).value = ""


         # ----- 3) En-têtes du tableau (ultra-robuste) -----
    def _norm(x: str) -> str:
        return _normalize_header_text(x)

    # cibles + synonymes acceptés
    order = ["Référence", "Produit", "DDM", "Quantité cartons", "Quantité palettes", "Poids palettes (kg)"]
    syns = {
        "Référence": ["référence", "reference"],
        "Produit": ["produit", "produit (gout + format)", "produit gout format"],
        "DDM": ["ddm", "date de durabilite", "date de durabilité"],
        "Quantité cartons": ["quantité cartons", "quantite cartons", "n° cartons", "no cartons", "nb cartons"],
        "Quantité palettes": ["quantité palettes", "quantite palettes", "n° palettes", "no palettes", "nb palettes"],
        "Poids palettes (kg)": ["poids palettes (kg)", "poids palettes", "poids (kg)"],
    }

    def _score_row(row_idx: int):
        max_cols = min(ws.max_column, 120)
        vals = [_norm(ws.cell(row=row_idx, column=j).value) for j in range(1, max_cols + 1)]
        cmap = {k: None for k in order}
        hits = 0
        for key in order:
            wanted = [_norm(key)] + [_norm(x) for x in syns[key]]
            # exact
            for j, hv in enumerate(vals, start=1):
                if hv in wanted:
                    cmap[key] = j
                    hits += 1
                    break
            # contains
            if cmap[key] is None:
                for j, hv in enumerate(vals, start=1):
                    if any(w in hv for w in wanted if len(w) >= 3):
                        cmap[key] = j
                        hits += 1
                        break
        return hits, cmap

    # 1) on scanne les premières lignes et on garde la meilleure
    best = (0, None, None)  # (hits, row, colmap)
    for r_scan in range(1, min(ws.max_row, 200) + 1):
        h, cmap = _score_row(r_scan)
        if h > best[0]:
            best = (h, r_scan, cmap)
        if h >= 4:
            break

    hits, hdr_row, colmap = best

    # 2) fallback ciblé : si <3 hits, cherche une cellule "référence" et rescore cette ligne
    if hits < 3:
        rr, cc = _find_cell_by_regex(ws, r"r[eé]f[ée]rence")
        if rr and cc:
            h2, cmap2 = _score_row(rr)
            if h2 >= hits:
                hits, hdr_row, colmap = h2, rr, cmap2

    if not hdr_row:
        raise KeyError("Ligne d’en-têtes du tableau introuvable dans le modèle Excel (pas assez de correspondances).")

    # 3) complète les colonnes manquantes en utilisant l'ordre gauche→droite de 'order'
    taken = {v for v in colmap.values() if v is not None}
    # point de départ : juste après "Référence" si on l'a, sinon colonne 1
    pos_guess = (colmap.get("Référence") or 1) + 1
    for key in order:
        if colmap.get(key) is None:
            # essaie de trouver une étiquette plausible autour
            found = None
            for j in range(pos_guess, min(ws.max_column, pos_guess + 12) + 1):
                if j in taken:
                    continue
                hv = _norm(ws.cell(row=hdr_row, column=j).value)
                wanted = [_norm(key)] + [_norm(x) for x in syns[key]]
                if hv in wanted or any(w in hv for w in wanted if len(w) >= 3):
                    found = j
                    break
            if found is None:
                # sinon, prend la prochaine colonne libre
                j = pos_guess
                while j in taken:
                    j += 1
                found = j
            colmap[key] = found
            taken.add(found)
            pos_guess = found + 1

    c_ref   = colmap["Référence"]
    c_prod  = colmap["Produit"]
    c_ddm   = colmap["DDM"]
    c_qc    = colmap["Quantité cartons"]
    c_qp    = colmap["Quantité palettes"]
    c_poids = colmap["Poids palettes (kg)"]

    # Force l'étiquette "Produit" visible dans le modèle
    _write_cell(ws, hdr_row, c_prod, "Produit")



    # ----- 4) Normalisation DF d'entrée -----
    df = df_lines.copy()
    # alias Produit
    if "Produit" not in df.columns and "Produit (goût + format)" in df.columns:
        df = df.rename(columns={"Produit (goût + format)": "Produit"})
    # DDM → texte jj/mm/aaaa
    def _to_ddm_val(x):
        if isinstance(x, (date, )):
            return x.strftime("%d/%m/%Y")
        s = str(x or "").strip()
        if not s:
            return ""
        # supports "yyyy-mm-dd" or "dd/mm/yyyy"
        try:
            if "-" in s and len(s.split("-")[0]) == 4:
                return datetime.strptime(s, "%Y-%m-%d").strftime("%d/%m/%Y")
            return datetime.strptime(s, "%d/%m/%Y").strftime("%d/%m/%Y")
        except Exception:
            return s

    # ----- 5) Écriture des lignes -----
    row = hdr_row + 1

    def _as_int(v) -> int:
        try:
            f = float(v)
            return int(round(f))
        except Exception:
            return 0

    for _, r in df.iterrows():
        ws.cell(row=row, column=c_ref).value   = str(r.get("Référence", ""))
        ws.cell(row=row, column=c_prod).value  = str(r.get("Produit", ""))
        ws.cell(row=row, column=c_ddm).value   = _to_ddm_val(r.get("DDM", ""))

        ws.cell(row=row, column=c_qc).value    = _as_int(r.get("Quantité cartons", 0))
        ws.cell(row=row, column=c_qp).value    = _as_int(r.get("Quantité palettes", 0))
        ws.cell(row=row, column=c_poids).value = _as_int(r.get("Poids palettes (kg)", 0))
        row += 1

    # ----- 6) Sauvegarde -----
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# =======================  PDF BL enlèvements (fpdf2)  =======================

def build_bl_enlevements_pdf(
    date_creation: date,
    date_ramasse: date,
    destinataire_title: str,
    destinataire_lines: List[str],
    df_lines: pd.DataFrame,
    *,
    logo_path: str | None = "assets/logo_symbiose.png",
    issuer_name: str = "FERMENT STATION",
    issuer_lines: List[str] | None = None,
    issuer_footer: str | None = "Produits issus de l'Agriculture Biologique certifié par FR-BIO-01",
) -> bytes:
    """PDF BL au look Excel : encadré, tableau gris, totaux. (Helvetica/latin-1)."""
    import os
    from fpdf import FPDF

    # ---------- helpers texte latin-1 ----------
    def _latin1_safe(s: str) -> str:
        s = str(s or "")
        repl = {"—": "-", "–": "-", "‒": "-", "’": "'", "‘": "'", "“": '"', "”": '"', "…": "...",
                "\u00A0": " ", "\u202F": " ", "\u2009": " ", "œ": "oe", "Œ": "OE", "€": "EUR"}
        for k, v in repl.items():
            s = s.replace(k, v)
        return s.encode("latin-1", "replace").decode("latin-1")

    def _txt(x) -> str:
        return _latin1_safe(x)

    # ---------- data ----------
    df = df_lines.copy()
    if "Produit" not in df.columns and "Produit (goût + format)" in df.columns:
        df = df.rename(columns={"Produit (goût + format)": "Produit"})

    def _ival(x):
        try:
            return int(round(float(x)))
        except Exception:
            return 0

    # ---------- PDF ----------
    pdf = FPDF("P", "mm", "A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    left, right = 15, 195
    width = right - left

    # ---- Logo + coordonnées expéditeur
    y = 18
    if logo_path and os.path.exists(logo_path):
        pdf.image(logo_path, x=left, y=y - 2, w=28)
        x_text = left + 34
    else:
        x_text = left

    pdf.set_xy(x_text, y)
    pdf.set_font("Helvetica", "B", 12); pdf.cell(0, 6, _txt(issuer_name), ln=1)
    pdf.set_x(x_text); pdf.set_font("Helvetica", "", 11)
    if issuer_lines is None:
        issuer_lines = [
            "Carré Ivry Bâtiment D2",
            "47 rue Ernest Renan",
            "94200 Ivry-sur-Seine - FRANCE",
            "Tél : 0967504647",
            "Site : https://www.symbiose-kefir.fr",
        ]
    for line in issuer_lines:
        pdf.set_x(x_text); pdf.cell(0, 5, _txt(line), ln=1)
    if issuer_footer:
        pdf.ln(2); pdf.set_x(x_text); pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 4, _txt(issuer_footer), ln=1)
    pdf.ln(2)

    # ---- Encadré "BON DE LIVRAISON"
    x_box, w_box = left, width * 0.70
    w_lbl, w_val = w_box * 0.55, w_box * 0.45
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_xy(x_box, pdf.get_y() + 2)
    pdf.cell(w_box, 8, _txt("BON DE LIVRAISON"), border=1, ln=1)

    pdf.set_font("Helvetica", "", 11)

    def _row_simple(label: str, value: str):
        pdf.set_x(x_box)
        pdf.cell(w_lbl, 8, _txt(label), border=1)
        pdf.cell(w_val, 8, _txt(value), border=1, ln=1, align="R")

    def _row_dest(label: str, title: str, lines: List[str]):
        val_text = "\n".join([title] + (lines or []))
        n_lines = len(pdf.multi_cell(w_val, 6, _txt(val_text), split_only=True)) or 1
        row_h = max(8, 6 * n_lines)
        y0 = pdf.get_y()
        pdf.set_xy(x_box, y0); pdf.cell(w_lbl, row_h, _txt(label), border=1)
        pdf.set_xy(x_box + w_lbl, y0); pdf.multi_cell(w_val, 6, _txt(val_text), border=1)
        pdf.set_xy(x_box, y0 + row_h)

    _row_simple("DATE DE CREATION :", date_creation.strftime("%d/%m/%Y"))
    _row_simple("DATE DE RAMASSE :", date_ramasse.strftime("%d/%m/%Y"))
    _row_dest("DESTINATAIRE :", destinataire_title, destinataire_lines)

    # ---- Tableau
    pdf.ln(6)
    pdf.set_fill_color(230, 230, 230)

    headers = [
        "Référence",
        "Produit",
        "DDM",
        "N° cartons",
        "N° palettes",
        "Poids (kg)",
    ]

    # Largeurs (somme 180) — Référence & DDM élargies
    widths_base = [30, 66, 26, 24, 22, 12]
    widths = widths_base[:]
    header_h = 8
    line_h = 6

    # Auto-ajustement des titres sur 1 ligne
    pdf.set_font("Helvetica", "B", 10)
    margin_mm = 2.5
    min_w = {0: 30.0, 1: 58.0, 2: 26.0, 3: 22.0, 4: 20.0, 5: 18.0}
    extra_needed = 0.0
    for j, h in enumerate(headers):
        if j == 1:  # on prend l'espace à Produit si besoin
            continue
        need = pdf.get_string_width(_txt(h)) + 2 * margin_mm
        new_w = max(widths[j], need, min_w.get(j, widths[j]))
        extra_needed += max(0.0, new_w - widths_base[j])
        widths[j] = new_w
    widths[1] = max(min_w[1], widths[1] - extra_needed)
    total = sum(widths)
    if total > 180.0:
        overflow = total - 180.0
        take = min(overflow, max(0.0, widths[1] - min_w[1]))
        widths[1] -= take; overflow -= take
        for j in (3, 4, 5, 0, 2):
            if overflow <= 0: break
            free = max(0.0, widths[j] - min_w[j])
            d = min(free, overflow)
            widths[j] -= d; overflow -= d

    # En-tête (1 ligne)
    x = left; y = pdf.get_y()
    for h, w in zip(headers, widths):
        pdf.set_xy(x, y); pdf.cell(w, header_h, _txt(h), border=1, align="C", fill=True); x += w
    pdf.set_xy(left, y + header_h)

    # Lignes
    pdf.set_font("Helvetica", "", 10)
    tot_cart = tot_pal = tot_poids = 0

    def _maybe_break(h):
        if pdf.will_page_break(h + header_h):
            pdf.add_page()
            pdf.set_fill_color(230, 230, 230)
            pdf.set_font("Helvetica", "B", 10)
            xh = left; yh = pdf.get_y()
            for hh, ww in zip(headers, widths):
                pdf.set_xy(xh, yh); pdf.cell(ww, header_h, _txt(hh), border=1, align="C", fill=True); xh += ww
            pdf.set_xy(left, yh + header_h)
            pdf.set_font("Helvetica", "", 10)

    for _, r in df.iterrows():
        ref = _txt(r.get("Référence", ""))
        prod = _txt(r.get("Produit", ""))
        ddm = _txt(r.get("DDM", ""))
        qc = _ival(r.get("N° cartons", r.get("Quantité cartons", 0)));   tot_cart += qc
        qp = _ival(r.get("N° palettes", r.get("Quantité palettes", 0)));  tot_pal  += qp
        po = _ival(r.get("Poids (kg)",  r.get("Poids palettes (kg)", 0))); tot_poids += po

        prod_lines = pdf.multi_cell(widths[1], line_h, prod, split_only=True)
        row_h = max(line_h, line_h * len(prod_lines))
        _maybe_break(row_h)

        xrow = left; yrow = pdf.get_y()
        pdf.set_xy(xrow, yrow); pdf.multi_cell(widths[0], row_h, ref, border=1, align="C"); xrow += widths[0]
        pdf.set_xy(xrow, yrow); pdf.multi_cell(widths[1], line_h, prod, border=1, align="L", max_line_height=line_h); xrow += widths[1]
        pdf.set_xy(xrow, yrow); pdf.multi_cell(widths[2], row_h, ddm, border=1, align="C"); xrow += widths[2]
        pdf.set_xy(xrow, yrow); pdf.multi_cell(widths[3], row_h, str(qc), border=1, align="C"); xrow += widths[3]
        pdf.set_xy(xrow, yrow); pdf.multi_cell(widths[4], row_h, str(qp), border=1, align="C"); xrow += widths[4]
        pdf.set_xy(xrow, yrow); pdf.multi_cell(widths[5], row_h, str(po), border=1, align="C")
        pdf.set_xy(left, yrow + row_h)

    # Totaux
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(widths[0] + widths[1] + widths[2], 8, _txt("Totaux"), border=1, align="R")
    pdf.cell(widths[3], 8, _txt(f"{tot_cart:,}".replace(",", " ")), border=1, align="C")
    pdf.cell(widths[4], 8, _txt(f"{tot_pal:,}".replace(",", " ")),  border=1, align="C")
    pdf.cell(widths[5], 8, _txt(f"{tot_poids:,}".replace(",", " ")), border=1, align="C")

    return bytes(pdf.output(dest="S"))
