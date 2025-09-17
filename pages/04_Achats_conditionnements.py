# pages/04_Achats_conditionnements.py
from __future__ import annotations
import io
import re
import unicodedata
from typing import Tuple, List, Dict

import numpy as np
import pandas as pd
import streamlit as st

from common.design import apply_theme, section, kpi
from core.optimizer import parse_stock, VOL_TOL  # formats 12x33 / 6x75 / 4x75

# ====================== THEME & CONTEXTE ======================
apply_theme("Achats — Conditionnements", "📦")
section("Prévision d’achats (conditionnements)", "📦")

# Besoin du fichier ventes déjà chargé dans l'accueil
if "df_raw" not in st.session_state or "window_days" not in st.session_state:
    st.warning("Va d’abord dans **Accueil** pour déposer l’Excel des ventes/stock, puis reviens ici.")
    st.stop()

df_raw = st.session_state.df_raw.copy()
window_days = float(st.session_state.window_days)

# ====================== OPTIONS (sidebar) ======================
with st.sidebar:
    st.header("Période à prévoir")
    horizon_j = st.number_input("Horizon (jours)", min_value=1, max_value=365, value=14, step=1)
    st.caption("Le besoin prévoit une consommation sur cet horizon à partir des ventes moyennes.")
    st.markdown("---")
    st.header("Options étiquettes")
    force_labels = st.checkbox("Étiquettes = 1 par bouteille (forcer si 'étiquette' dans le nom)", value=True)

st.caption(
    f"Excel ventes courant : **{st.session_state.get('file_name','(sans nom)')}** — "
    f"Fenêtre de calcul des vitesses : **{int(window_days)} jours** — "
    f"Horizon prévision : **{int(horizon_j)} jours**"
)

# ====================== IMPORTS FICHIERS (dans la page) ======================
section("Importer les fichiers", "📥")
c1, c2 = st.columns(2)
with c1:
    st.subheader("Consommation des articles (Excel)")
    conso_file = st.file_uploader(
        "Déposer le fichier *Consommation* ici",
        type=["xlsx","xls"],
        key="uploader_conso",
        label_visibility="collapsed"
    )
with c2:
    st.subheader("Stocks des articles (Excel)")
    stock_file = st.file_uploader(
        "Déposer le fichier *Stocks* ici",
        type=["xlsx","xls"],
        key="uploader_stock",
        label_visibility="collapsed"
    )

# ====================== HELPERS TEXTE ======================
def _canon_txt(s: str) -> str:
    """Minuscule, sans accents, espaces→1, supprime ponctuation superflue."""
    s = str(s or "")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^a-zA-Z0-9]+", " ", s).strip().lower()
    return s

def _is_total_row(s: str) -> bool:
    """Détecte 'TOTAL', 'Total général', etc. (robuste aux accents/casse)."""
    t = _canon_txt(s)
    if not t:
        return False
    if t.startswith("total"):
        return True
    return t in {
        "total general", "grand total", "totaux", "total stock",
        "total stocks", "total consommation", "total consommations",
        "total achats", "total des achats"
    }

# ====================== LECTURE FICHIERS ======================
@st.cache_data(show_spinner=False)
def read_consumption_xlsx(uploaded_consumption_file) -> pd.DataFrame:
    """
    Lit la **colonne B** depuis la ligne **juste après** la cellule contenant 'conditionnement'
    jusqu’à **2 lignes avant** la cellule contenant 'contenants'.
    Chaque cellule de la colonne B contient un libellé + un nombre (coeff. de conso).
    Renvoie colonnes: key | article | conso | per_hint ('bottle'/'carton')
    - Lignes TOTAL ignorées.
    - Si 'étiquette' dans le nom: per_hint='bottle' et conso=1.0 (sera renforcé côté calcul si force_labels).
    """
    import openpyxl

    bio = io.BytesIO(uploaded_consumption_file.read())
    wb = openpyxl.load_workbook(bio, data_only=True)
    ws = wb[wb.sheetnames[0]]

    start_row, end_row = None, None
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column, 20) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                t = _canon_txt(v)
                if start_row is None and "conditionnement" in t:
                    start_row = r + 1  # lecture à partir de la ligne suivante, col B
                if "contenants" in t:
                    end_row = (r - 2) if start_row else None  # 2 lignes avant
    if start_row is None or end_row is None or end_row < start_row:
        return pd.DataFrame(columns=["key", "article", "conso", "per_hint"])

    data = []
    for r in range(start_row, end_row + 1):
        raw = ws.cell(row=r, column=2).value  # colonne B
        if raw is None:
            continue
        key = str(raw).strip()
        if key == "" or _is_total_row(key):
            continue

        # Extrait le premier nombre comme coefficient (ex: "Capsules 26 - 1" -> 1)
        s = str(key)
        m = re.search(r"([0-9]+(?:[.,][0-9]+)?)", s)
        if m:
            conso = float(m.group(1).replace(",", "."))
            article = (s[:m.start()] + s[m.end():]).strip(" -–—:;/\t")
        else:
            conso = 1.0
            article = s

        # Heuristique d'unité
        per_hint = "bottle"
        if re.search(r"\b(carton|caisse|colis|etui|étui|pack|boite|boîte)\b", article, flags=re.I):
            per_hint = "carton"

        # Cas étiquette: forcer "par bouteille" (+ conso=1.0 par défaut)
        if re.search(r"etiquette|étiquette", article, flags=re.I):
            per_hint = "bottle"
            # on garde conso détecté; il pourra être forcé à 1.0 au calcul si option activée

        data.append({
            "key": _canon_txt(article),
            "article": article.strip(),
            "conso": float(conso),
            "per_hint": per_hint,
        })

    block = pd.DataFrame(data)
    if block.empty:
        return block
    # Agrège doublons + filtre TOTAL de sécurité
    block = block[~block["article"].apply(_is_total_row)]
    block = (block.groupby(["key", "article", "per_hint"], as_index=False)["conso"].sum())
    return block.reset_index(drop=True)

@st.cache_data(show_spinner=False)
def read_stock_xlsx(uploaded_stocks_file) -> pd.DataFrame:
    """
    Lit l’Excel Stocks et récupère:
      - 'article' (libellé)
      - 'stock' depuis la colonne 'Quantité virtuelle' (ou colonne F si non trouvée)
    Filtre les lignes TOTAL et agrège les doublons.
    Renvoie: key | article | stock
    """
    # 1) tentative avec en-têtes nommés
    try:
        bio = io.BytesIO(uploaded_stocks_file.read())
        df = pd.read_excel(bio, engine="openpyxl")
    except Exception:
        bio = io.BytesIO(uploaded_stocks_file.read())
        df = pd.read_excel(bio)

    cols_lower = {str(c).strip().lower(): c for c in df.columns}
    qty_col = None
    for k, real in cols_lower.items():
        if ("quantité virtuelle" in k) or ("quantite virtuelle" in k) or ("qte virtuelle" in k):
            qty_col = real; break

    # colonne article
    name_col = None
    for cand in ["article","libellé","libelle","désignation","designation","intitulé","intitule","nom"]:
        if cand in cols_lower:
            name_col = cols_lower[cand]; break
    if name_col is None and len(df.columns):
        name_col = df.columns[0]

    if qty_col is None:
        # 2) fallback: considère la 6e colonne (F) si disponible, sans en-tête fiable
        df0 = df.copy()
        if df0.shape[1] >= 6:
            qty_col = df0.columns[5]
        else:
            # rien d'exploitable
            return pd.DataFrame(columns=["key","article","stock"])

    out = df[[name_col, qty_col]].copy()
    out.columns = ["article", "stock"]
    out["article"] = out["article"].astype(str).str.strip()
    out = out[~out["article"].apply(_is_total_row)]
    out["stock"] = pd.to_numeric(out["stock"], errors="coerce").fillna(0.0)
    out["key"] = out["article"].map(_canon_txt)
    out = out.groupby(["key","article"], as_index=False)["stock"].sum()
    return out.reset_index(drop=True)

# ====================== CALCULS VENTES → FORMATS ======================
def _fmt_from_stock_text(stock_txt: str) -> str | None:
    """Retourne '12x33' / '6x75' / '4x75' depuis la colonne Stock."""
    nb, vol = parse_stock(stock_txt)
    if pd.isna(nb) or pd.isna(vol): return None
    nb = int(nb); vol = float(vol)
    if nb == 12 and abs(vol - 0.33) <= VOL_TOL: return "12x33"
    if nb == 6  and abs(vol - 0.75) <= VOL_TOL: return "6x75"
    if nb == 4  and abs(vol - 0.75) <= VOL_TOL: return "4x75"
    return None

def aggregate_forecast_by_format(df_sales: pd.DataFrame, window_days: float, horizon_j: int) -> Dict[str, Dict[str, float]]:
    """Calcule bouteilles et cartons *prévisionnels* par format sur l’horizon H à partir des vitesses moyennes."""
    req = ["Stock", "Volume vendu (hl)"]
    if any(c not in df_sales.columns for c in req):
        return {}

    tmp = df_sales.copy()
    tmp["fmt"] = tmp["Stock"].map(_fmt_from_stock_text)
    tmp = tmp.dropna(subset=["fmt"])

    parsed = tmp["Stock"].map(parse_stock)
    tmp[["nb_btl_cart", "vol_L"]] = pd.DataFrame(parsed.tolist(), index=tmp.index)
    tmp["vol_hL_per_btl"] = (pd.to_numeric(tmp["vol_L"], errors="coerce") / 100.0)
    tmp["nb_btl_cart"] = pd.to_numeric(tmp["nb_btl_cart"], errors="coerce")

    tmp["v_hL_j"] = pd.to_numeric(tmp["Volume vendu (hl)"], errors="coerce") / max(float(window_days), 1.0)
    tmp = tmp.replace([np.inf, -np.inf], np.nan).dropna(subset=["vol_hL_per_btl", "nb_btl_cart", "v_hL_j"])

    tmp["btl_j"] = np.where(tmp["vol_hL_per_btl"] > 0, tmp["v_hL_j"] / tmp["vol_hL_per_btl"], 0.0)
    tmp["carton_j"] = np.where(tmp["nb_btl_cart"] > 0, tmp["btl_j"] / tmp["nb_btl_cart"], 0.0)

    tmp["btl_h"] = horizon_j * tmp["btl_j"]
    tmp["carton_h"] = horizon_j * tmp["carton_j"]

    agg = tmp.groupby("fmt").agg(bottles=("btl_h","sum"), cartons=("carton_h","sum"))
    out = {fmt: {"bottles": float(agg.loc[fmt, "bottles"]), "cartons": float(agg.loc[fmt, "cartons"])} for fmt in agg.index}
    for k in ["12x33", "6x75", "4x75"]:
        out.setdefault(k, {"bottles": 0.0, "cartons": 0.0})
    return out

# ====================== MAPPING ARTICLES → FORMATS ======================
def _article_applies_formats(article: str) -> Tuple[List[str], str]:
    """
    Formats cibles + unité par défaut.
    - '12x33' → 12x33 ; '6x75' → 6x75 ; '4x75' → 4x75
    - '33' seul → 12x33 ; '75' → 6x75 & 4x75 ; sinon → tous formats
    - 'carton/caisse/colis/étui/pack/boîte' → unité 'carton', sinon 'bottle'
    """
    a = _canon_txt(article)
    per = "carton" if any(w in a for w in ["carton","caisse","colis","etui","étui","pack","boite","boîte"]) else "bottle"
    if "12x33" in a: fmts = ["12x33"]
    elif "6x75" in a: fmts = ["6x75"]
    elif "4x75" in a: fmts = ["4x75"]
    elif "33" in a and "75" not in a: fmts = ["12x33"]
    elif "75" in a: fmts = ["6x75","4x75"]
    else: fmts = ["12x33","6x75","4x75"]
    return fmts, per

# ====================== BESOIN & ACHATS ======================
def compute_needs_table(
    df_conso: pd.DataFrame,
    df_stock: pd.DataFrame,
    forecast_fmt: Dict[str, Dict[str, float]],
    *,
    force_labels: bool
) -> pd.DataFrame:
    """
    Besoin = conso × (bouteilles ou cartons prévus) agrégé par formats applicables.
    Cas particulier ÉTIQUETTES (option):
      - si 'étiquette' dans le nom ET option cochée → conso = 1 par bouteille
    Fusionne ensuite avec le stock (clé = libellé canonisé) et calcule À acheter.
    """
    rows = []
    for _, r in df_conso.iterrows():
        art = str(r["article"]).strip()
        key = _canon_txt(art)
        conso = float(r["conso"])
        fmts, per = _article_applies_formats(art)

        # Règle spéciale étiquettes
        if force_labels and re.search(r"\b(etiquette|étiquette|etiquettes|étiquettes)\b", art, flags=re.I):
            per = "bottle"
            conso = 1.0  # 1 étiquette par bouteille

        # sinon on respecte le per_hint lu dans le fichier conso
        ph = str(r.get("per_hint", "")).strip().lower()
        if ph in ("bottle", "carton"):
            per = ph

        qty = 0.0
        for f in fmts:
            if per == "bottle":
                qty += conso * float(forecast_fmt.get(f, {}).get("bottles", 0.0))
            else:
                qty += conso * float(forecast_fmt.get(f, {}).get("cartons", 0.0))

        rows.append({
            "key": key,
            "Article": art,
            "Unité": "par bouteille" if per == "bottle" else "par carton",
            "Besoin horizon": qty
        })

    need_df = pd.DataFrame(rows)
    if need_df.empty:
        return pd.DataFrame(columns=["Article","Unité","Besoin horizon","Stock dispo","À acheter"])

    st_df = (
        df_stock[["key","stock"]].rename(columns={"stock":"Stock dispo"})
        if (df_stock is not None and not df_stock.empty)
        else pd.DataFrame(columns=["key","Stock dispo"])
    )

    out = need_df.merge(st_df, on="key", how="left").fillna({"Stock dispo": 0.0})

    # À acheter = max(Besoin - Stock, 0) puis arrondi entier
    out["À acheter"] = np.maximum(out["Besoin horizon"] - out["Stock dispo"], 0.0)

    # Arrondis (entiers à l’affichage et à l’export)
    for c in ["Besoin horizon", "Stock dispo", "À acheter"]:
        out[c] = np.round(out[c], 0).astype(int)

    return out.drop(columns=["key"]).sort_values("À acheter", ascending=False).reset_index(drop=True)

# ====================== CALCULS GLOBAUX ======================
forecast = aggregate_forecast_by_format(df_raw, window_days=window_days, horizon_j=int(horizon_j))

# KPIs — afficher des ÉTIQUETTES (≈ nb de bouteilles)
b_33 = float(forecast.get("12x33", {}).get("bottles", 0.0))
b_75 = float(forecast.get("6x75", {}).get("bottles", 0.0) + forecast.get("4x75", {}).get("bottles", 0.0))
cartons_total = float(sum(v.get("cartons", 0.0) for v in forecast.values()))

colA, colB, colC = st.columns([1.1, 1, 1])
with colA:
    kpi("Étiquettes à prévoir — 12x33", f"{b_33:.0f}")
with colB:
    kpi("Étiquettes à prévoir — 75cl", f"{b_75:.0f}")
with colC:
    kpi("Cartons prévus (tous formats)", f"{cartons_total:.0f}")

# ====================== LECTURE + RÉSULTAT ======================
df_conso = None
df_stockc = None
err_block = False

if conso_file is not None:
    try:
        df_conso = read_consumption_xlsx(conso_file)
        st.success("Consommation: zone **colonne B** détectée ✅")
        with st.expander("Voir l’aperçu du fichier **Consommation**", expanded=False):
            st.dataframe(df_conso[["article","conso","per_hint"]], use_container_width=True, hide_index=True)
    except Exception as e:
        st.error(f"Erreur lecture consommation: {e}")
        err_block = True
else:
    st.info("Importer l’Excel **Consommation des articles** (bloc ci-dessus).")

if stock_file is not None:
    try:
        df_stockc = read_stock_xlsx(stock_file)
        st.success("Stocks: colonne **Quantité virtuelle** détectée (ou F) ✅")
        with st.expander("Voir l’aperçu du fichier **Stocks**", expanded=False):
            st.dataframe(df_stockc[["article","stock"]], use_container_width=True, hide_index=True)
    except Exception as e:
        st.error(f"Erreur lecture stocks: {e}")
        err_block = True
else:
    st.info("Importer l’Excel **Stocks des articles** (bloc ci-dessus).")

st.markdown("---")

if (df_conso is not None) and (df_stockc is not None) and (not err_block):
    result = compute_needs_table(df_conso, df_stockc, forecast, force_labels=force_labels)

    if result.empty:
        st.info("Aucun besoin calculé (vérifie les fichiers de consommation/stocks et les correspondances d’articles).")
        st.stop()

    total_buy = int(result["À acheter"].sum())
    nb_items  = int((result["À acheter"] > 0).sum())
    c1, c2 = st.columns(2)
    with c1: kpi("Articles à acheter (nb)", f"{nb_items}")
    with c2: kpi("Quantité totale à acheter (unités)", f"{total_buy:,}".replace(",", " "))

    st.subheader("Proposition d’achats (triée par 'À acheter' décroissant)")
    st.dataframe(
        result[["Article","Unité","Besoin horizon","Stock dispo","À acheter"]],
        use_container_width=True, hide_index=True,
        column_config={
            "Besoin horizon": st.column_config.NumberColumn(format="%d"),
            "Stock dispo": st.column_config.NumberColumn(format="%d"),
            "À acheter": st.column_config.NumberColumn(format="%d"),
        }
    )

    csv_bytes = result.to_csv(index=False).encode("utf-8")
    st.download_button(
        "⬇️ Exporter la proposition (CSV)",
        data=csv_bytes,
        file_name=f"achats_conditionnements_{int(horizon_j)}j.csv",
        mime="text/csv",
        use_container_width=True,
    )
else:
    st.info("Charge les deux fichiers pour obtenir la proposition d’achats.")
