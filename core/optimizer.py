
- Fichier = app.py : 
# app.py — Accueil + préflight syntaxe des pages
import pathlib, traceback
import streamlit as st
import pandas as pd

# ---------- PRE-FLIGHT : détecte les erreurs de syntaxe dans pages/*.py ----------
def _preflight_pages():
    root = pathlib.Path(__file__).resolve().parent
    pages = sorted((root / "pages").glob("*.py"))
    bad = []
    for p in pages:
        code = p.read_text(encoding="utf-8", errors="replace")
        try:
            compile(code, str(p), "exec")
        except SyntaxError as e:
            st.set_page_config(page_title="Erreur de syntaxe", page_icon="🛑", layout="wide")
            st.title("🛑 Erreur de syntaxe dans une page Streamlit")
            st.error(f"Fichier : `{p.name}` — ligne **{e.lineno}**, colonne **{e.offset}**")
            st.code("".join(traceback.format_exception_only(e)), language="text")
            # extrait de code : 2 lignes avant/après
            lines = code.splitlines()
            i = max(0, (e.lineno or 1) - 1)
            snippet = "\n".join(lines[max(0, i-2): i+3])
            st.code(snippet, language="python")
            st.info("Corrige ce fichier dans GitHub → Commit → recharge l’app.")
            bad.append(p)
    if bad:
        st.stop()

_preflight_pages()
# ---------- FIN PRE-FLIGHT ------------------------------------------------------

# --- Accueil “Uploader unique” (ton code d’origine) ---
from common.design import apply_theme, section
from core.optimizer import read_input_excel_and_period_from_upload

apply_theme("Ferment Station — Accueil", "🥤")
section("Accueil", "🏠")
st.caption("Dépose ici ton fichier Excel. Il sera utilisé automatiquement dans tous les onglets.")

uploaded = st.file_uploader("Dépose un Excel (.xlsx / .xls)", type=["xlsx", "xls"])
col1, col2 = st.columns([1,1])
with col1:
    clear = st.button("♻️ Réinitialiser le fichier chargé", use_container_width=True)
with col2:
    show_head = st.toggle("Afficher un aperçu (20 premières lignes)", value=True)

if clear:
    for k in ("df_raw", "window_days", "file_name"):
        if k in st.session_state:
            del st.session_state[k]
    st.success("Fichier déchargé. Dépose un nouvel Excel pour continuer.")

if uploaded is not None:
    try:
        df_raw, window_days = read_input_excel_and_period_from_upload(uploaded)
        st.session_state.df_raw = df_raw
        st.session_state.window_days = window_days
        st.session_state.file_name = uploaded.name
        st.success(f"Fichier chargé ✅ : **{uploaded.name}** · Fenêtre détectée (B2) : **{window_days} jours**")
    except Exception as e:
        st.error(f"Erreur de lecture de l'Excel : {e}")

if "df_raw" in st.session_state:
    st.info(f"Fichier en mémoire : **{st.session_state.get('file_name','(sans nom)')}** — fenêtre : **{st.session_state.get('window_days', '—')} jours**")
    if show_head:
        st.dataframe(st.session_state.df_raw.head(20), use_container_width=True)
else:
    st.warning("Aucun fichier en mémoire. Dépose un Excel ci-dessus pour activer les autres onglets.")

- Fichier = README.md : 
# Ferment Station — Streamlit (multi-pages)

L'app lit **uniquement** les fichiers du repo (`/data`, `/assets`).  
Aucune importation locale n'est nécessaire.

## Structure
- `app.py` (accueil)
- `pages/01_Production.py`, `pages/02_Optimisation.py`, `pages/03_Fiche_de_ramasse.py`
- `common/design.py` (thème & UI)
- `common/data.py` (config & chemins)
- `core/optimizer.py` (algorithmes)
- `data/production.xlsx`, `data/flavor_map.csv`
- `assets/` (images produits)

## Lancer en local (optionnel)
```bash
pip install -r requirements.txt
streamlit run app.py

- Fichier = config.yaml : 
data_files:
  main_table: "data/production.xlsx"   # table source
  flavor_map: "data/flavor_map.csv"    # mapping nom → goût canonique
images_dir: "assets"                   # dossier des visuels produit

- Fichier = info_FDR.csv : 
Produit,Format,Désignation,Contenant,Packaging,Code-barre,Poids
"Kéfir de fruits Original","6x75cl","KÉFIR DE FRUITS ORIGINAL - 6X75CL (3383)","Bouteille 75cl EAU GAZEUSE - 0.75L","Carton de 6","3770014427052",7.23
"Kéfir de fruits Original","12x33cl","KÉFIR DE FRUITS ORIGINAL - 12X33CL (12)","Bouteille - 0.33L","Carton de 12","3770014427007",7.56
"Kéfir Gingembre","4x75cl","KÉFIR GINGEMBRE - 4X75CL (3382)","Bouteille 75cl SAFT - 0.75L","Pack de 4","23770014427049",4.68
"Kéfir Gingembre","6x75cl","KÉFIR GINGEMBRE - 6X75CL (3383)","Bouteille 75cl EAU GAZEUSE - 0.75L","Carton de 6","3770014427045",7.23
"Kéfir Gingembre","12x33cl","KÉFIR GINGEMBRE - 12X33CL (12)","Bouteille - 0.33L","Carton de 12","3770014427014",7.56
"Kéfir Mangue Passion","4x75cl","KÉFIR MANGUE PASSION - 4X75CL (3382)","Bouteille 75cl SAFT - 0.75L","Pack de 4","23770014427193",4.68
"Kéfir Mangue Passion","12x33cl","KÉFIR MANGUE PASSION - 12X33CL (12)","Bouteille - 0.33L","Carton de 12","3770014427038",7.56
"Kéfir Mangue Passion","6x75cl","KÉFIR MANGUE PASSION - 6X75CL (3383)","Bouteille 75cl EAU GAZEUSE - 0.75L","Carton de 6","3770014427199",7.23
"Kéfir menthe citron vert","4x75cl","KÉFIR MENTHE CITRON VERT - 4X75CL (3382)","Bouteille 75cl SAFT - 0.75L","Pack de 4","23770014427063",4.68
"Kéfir menthe citron vert","12x33cl","KÉFIR MENTHE CITRON VERT - 12X33CL (12)","Bouteille - 0.33L","Carton de 12","3770014427076",7.56
"Kéfir menthe citron vert","6x75cl","KÉFIR MENTHE CITRON VERT - 6X75CL (3383)","Bouteille 75cl EAU GAZEUSE - 0.75L","Carton de 6","3770014427069",7.23
"Infusion probiotique menthe poivrée","12x33cl","INFUSION PROBIOTIQUE MENTHE POIVRÉE - 12X33CL (12)","Bouteille - 0.33L","Carton de 12","3770014427182",7.56
"Kéfir Pamplemousse","4x75cl","KÉFIR PAMPLEMOUSSE - 4X75CL (3382)","Bouteille 75cl SAFT - 0.75L","Pack de 4","23770014427254",4.68
"Kéfir Pamplemousse","6x75cl","KÉFIR PAMPLEMOUSSE - 6X75CL (3383)","Bouteille 75cl EAU GAZEUSE - 0.75L","Carton de 6","3770014427250",7.23
"Kéfir Pamplemousse","12x33cl","KÉFIR PAMPLEMOUSSE - 12X33CL (12)","Bouteille - 0.33L","Carton de 12","3770014427267",7.56
"Infusion probiotique Anis","12x33cl","INFUSION PROBIOTIQUE ANIS - 12X33CL (12)","Bouteille - 0.33L","Carton de 12","3770014427175",7.56
"IGEBA Pêche","12x33cl","IGEBA PÊCHE - 12X33CL (12)","Bouteille - 0.33L","Carton de 12","3770014427274",7.56
"Infusion probiotique Mélisse","12x33cl","INFUSION PROBIOTIQUE MÉLISSE - 12X33CL (12)","Bouteille - 0.33L","Carton de 12","3770014427168",7.56
"Infusion probiotique Zest d'agrumes","12x33cl","INFUSION PROBIOTIQUE ZEST D'AGRUMES - 12X33CL (12)","Bouteille - 0.33L","Carton de 12","3770014427304",7.56
"Probiotic water Lemonbalm","12x33cl","INTER - INFUSION PROBIOTIQUE MELISSE 33 CL* (X12) (12)","Bouteille - 0.33L","Carton de 12","3770014427236",7.56
"Probiotic water Peppermint","12x33cl","INTER - INFUSION PROBIOTIQUE MENTHE POIVREE 33 CL* (12)","Bouteille - 0.33L","Carton de 12","3770014427229",7.56
"Water kefir Mango Passion","12x33cl","INTER - KEFIR MANGUE PASSION 33 CL* (X12) (12)","Bouteille - 0.33L","Carton de 12","3770014427021",7.56
"Water kefir Mint Lime","12x33cl","INTER - KEFIR MENTHE CITRON VERT 33 CL* (X12) (12)","Bouteille - 0.33L","Carton de 12","3770014427342",7.56
"Water kefir Grapefruit","12x33cl","INTER - KEFIR PAMPLEMOUSSE 33 CL* (X12) (12)","Bouteille - 0.33L","Carton de 12","3770014427243",6.741
"NIKO - Kéfir de fruits Menthe citron vert","12x33cl","NIKO - KEFIR MENTHE CITRON VERT 33 CL* (X12) (12)","Bouteille - 0.33L","Carton de 12","13770014427363",6.741
"NIKO - Kéfir de fruits Mangue Passion","12x33cl","NIKO - KEFIR MANGUE 33 CL* (X12) (12)","Bouteille - 0.33L","Carton de 12","137700144271?","6.741"
"NIKO - Kéfir de fruits Gingembre","12x33cl","NIKO - KEFIR GINGEMBRE 33 CL* (X12) (12)","Bouteille - 0.33L","Carton de 12","13770014427??","6.741"

- Fichier = requirements.txt : 
streamlit
pandas
numpy
openpyxl
pillow
pyyaml
xlrd
reportlab
fpdf2
python-dateutil
pdfplumber
pypdf
tomli

-DOSSIER 1 : pages 
-Fichier = 01_Production.py :
# pages/01_Production.py
import os
import re
import datetime as _dt
import pandas as pd
import streamlit as st
from dateutil.relativedelta import relativedelta

from common.design import apply_theme, section, kpi, find_image_path, load_image_bytes
from common.data import get_paths
from core.optimizer import (
    load_flavor_map_from_path,
    apply_canonical_flavor, sanitize_gouts,
    compute_plan,
)
from common.xlsx_fill import fill_fiche_7000L_xlsx

# ====== Réglages modèle Excel ======
TEMPLATE_PATH = "assets/Fiche de Prod 250620.xlsx"
SHEET_NAME = None

# ---------------- UI header ----------------
apply_theme("Production — Ferment Station", "📦")
section("Tableau de production", "📦")

# ---------------- Pré-requis : fichier chargé sur Accueil ----------------
if "df_raw" not in st.session_state or "window_days" not in st.session_state:
    st.warning("Aucun fichier chargé. Va dans **Accueil** pour déposer l'Excel, puis reviens.")
    st.stop()

# chemins (repo)
_, flavor_map, images_dir = get_paths()

# Données depuis l'accueil
df_in_raw = st.session_state.df_raw
window_days = st.session_state.window_days

# ---------------- Préparation des données ----------------
fm = load_flavor_map_from_path(flavor_map)
try:
    df_in = apply_canonical_flavor(df_in_raw, fm)
except KeyError as e:
    st.error(f"{e}")
    st.info("Astuce : vérifie la 1ère ligne (en-têtes) de ton Excel et renomme la colonne du nom produit en **'Produit'** ou **'Désignation'**.")
    st.stop()

df_in["Produit"] = df_in["Produit"].astype(str)
df_in = sanitize_gouts(df_in)

# ---------------- Sidebar (paramètres) ----------------
with st.sidebar:
    st.header("Paramètres")
    volume_cible = st.number_input("Volume cible (hL)", 1.0, 1000.0, 64.0, 1.0)
    nb_gouts = st.selectbox("Nombre de goûts simultanés", [1, 2], index=0)
    repartir_pro_rv = st.checkbox("Répartition au prorata des ventes", value=True)

    st.markdown("---")
    st.subheader("Filtres")
    all_gouts = sorted(pd.Series(df_in.get("GoutCanon", pd.Series(dtype=str))).dropna().astype(str).str.strip().unique())
    excluded_gouts = st.multiselect("🚫 Exclure certains goûts", options=all_gouts, default=[])

    # 🔥 NOUVEAU : forcer certains goûts
    forced_gouts = st.multiselect(
        "✅ Forcer la production de ces goûts",
        options=[g for g in all_gouts if g not in set(excluded_gouts)],
        help="Les goûts sélectionnés ici seront produits quoi qu’il arrive. "
             "Si tu en choisis plus que le nombre de goûts sélectionnés ci-dessus, "
             "le nombre sera automatiquement augmenté."
    )


st.caption(
    f"Fichier courant : **{st.session_state.get('file_name','(sans nom)')}** — Fenêtre (B2) : **{window_days} jours**"
)

# ---------------- Calculs ----------------
# Nombre de goûts effectif : on garantit que tous les 'forcés' rentrent
effective_nb_gouts = max(nb_gouts, len(forced_gouts)) if forced_gouts else nb_gouts

df_min, cap_resume, gouts_cibles, synth_sel, df_calc, df_all = compute_plan(
    df_in=df_in,
    window_days=window_days,
    volume_cible=volume_cible,
    nb_gouts=effective_nb_gouts,         # 👈 prend en compte les 'forcés'
    repartir_pro_rv=repartir_pro_rv,
    manual_keep=forced_gouts or None,    # 👈 forçage
    exclude_list=excluded_gouts,
)

# ---------------- KPIs ----------------
total_btl = int(pd.to_numeric(df_min.get("Bouteilles à produire (arrondi)"), errors="coerce").fillna(0).sum()) if "Bouteilles à produire (arrondi)" in df_min.columns else 0
total_vol = float(pd.to_numeric(df_min.get("Volume produit arrondi (hL)"), errors="coerce").fillna(0).sum()) if "Volume produit arrondi (hL)" in df_min.columns else 0.0
c1, c2, c3 = st.columns(3)
with c1: kpi("Total bouteilles à produire", f"{total_btl:,}".replace(",", " "))
with c2: kpi("Volume total (hL)", f"{total_vol:.2f}")
with c3: kpi("Goûts sélectionnés", f"{len(gouts_cibles)}")

# ---------------- Images + tableau principal ----------------
def sku_guess(name: str):
    m = re.search(r"\b([A-Z]{3,6}-\d{2,3})\b", str(name))
    return m.group(1) if m else None

df_view = df_min.copy()
df_view["SKU?"] = df_view["Produit"].apply(sku_guess)
df_view["__img_path"] = [
    find_image_path(images_dir, sku=sku_guess(p), flavor=g)
    for p, g in zip(df_view["Produit"], df_view["GoutCanon"])
]
df_view["Image"] = df_view["__img_path"].apply(load_image_bytes)

st.data_editor(
    df_view[[
        "Image","GoutCanon","Produit","Stock",
        "Cartons à produire (arrondi)","Bouteilles à produire (arrondi)",
        "Volume produit arrondi (hL)"
    ]],
    use_container_width=True,
    hide_index=True,
    disabled=True,
    column_config={
        "Image": st.column_config.ImageColumn("Image", width="small"),
        "GoutCanon": "Goût",
        "Volume produit arrondi (hL)": st.column_config.NumberColumn(format="%.2f"),
    },
)

# ======================================================================
# ========== Sauvegarde + génération de la fiche Excel ==================
# ======================================================================
section("Fiche de production (modèle Excel)", "🧾")

_sp_prev = st.session_state.get("saved_production")
default_debut = _dt.date.fromisoformat(_sp_prev["semaine_du"]) if _sp_prev and "semaine_du" in _sp_prev else _dt.date.today()

# Champ unique : date de début fermentation
date_debut = st.date_input("Date de début de fermentation", value=default_debut)

# DDM = début + 1 an
date_ddm = date_debut + _dt.timedelta(days=365)


if st.button("💾 Sauvegarder cette production", use_container_width=True):
    g_order = []
    if isinstance(df_min, pd.DataFrame) and "GoutCanon" in df_min.columns:
        for g in df_min["GoutCanon"].astype(str).tolist():
            if g and g not in g_order:
                g_order.append(g)

    st.session_state.saved_production = {
        "df_min": df_min.copy(),
        "df_calc": df_calc.copy(),
        "gouts": g_order,
        "semaine_du": date_debut.isoformat(),   # renommé mais même logique
        "ddm": date_ddm.isoformat(),
    }

    st.success("Production sauvegardée ✅ — tu peux maintenant générer la fiche.")

sp = st.session_state.get("saved_production")

def _two_gouts_auto(sp_obj, df_min_cur, gouts_cur):
    if isinstance(sp_obj, dict):
        g_saved = sp_obj.get("gouts")
        if g_saved:
            uniq = []
            for g in g_saved:
                if g and g not in uniq:
                    uniq.append(g)
            if uniq:
                return (uniq + [None, None])[:2]
    if isinstance(df_min_cur, pd.DataFrame) and "GoutCanon" in df_min_cur.columns:
        seen = []
        for g in df_min_cur["GoutCanon"].astype(str).tolist():
            if g and g not in seen:
                seen.append(g)
        if seen:
            return (seen + [None, None])[:2]
    base = list(gouts_cur) if gouts_cur else []
    return (base + [None, None])[:2]

if sp:
    g1, g2 = _two_gouts_auto(sp, sp.get("df_min", df_min), gouts_cibles)

    if not os.path.exists(TEMPLATE_PATH):
        st.error(f"Modèle introuvable. Place le fichier **{TEMPLATE_PATH}** dans le repo.")
    else:
        try:
            xlsx_bytes = fill_fiche_7000L_xlsx(
                template_path=TEMPLATE_PATH,
                semaine_du=_dt.date.fromisoformat(sp["semaine_du"]),
                ddm=_dt.date.fromisoformat(sp["ddm"]),
                gout1=g1 or "",
                gout2=g2,
                df_calc=sp.get("df_calc", df_calc),
                sheet_name=SHEET_NAME,
                df_min=sp.get("df_min", df_min),
            )

            semaine_label = _dt.date.fromisoformat(sp["semaine_du"]).strftime("%d-%m-%Y")
            fname_xlsx = f"Fiche de production (semaine du {semaine_label}).xlsx"

            st.download_button(
                "📄 Télécharger la fiche (XLSX, 2 pages, identique au modèle)",
                data=xlsx_bytes,
                file_name=fname_xlsx,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except FileNotFoundError:
            st.error("Modèle introuvable. Vérifie le chemin du fichier modèle.")
        except Exception as e:
            st.error(f"Erreur lors du remplissage du modèle : {e}")
else:
    st.info("Sauvegarde la production ci-dessus pour activer la génération de la fiche.")

-Fichier = 02.Optimisation.py :
# pages/02_Optimisation.py
import pandas as pd
import streamlit as st

from common.design import apply_theme, section, kpi
from common.data import get_paths
from core.optimizer import (
    load_flavor_map_from_path,
    apply_canonical_flavor,
    compute_losses_table_v48,
)

apply_theme("Optimisation & pertes — Ferment Station", "📉")
section("Optimisation & pertes", "📉")

# Besoin du fichier en mémoire
if "df_raw" not in st.session_state or "window_days" not in st.session_state:
    st.warning("Aucun fichier chargé. Va dans **Accueil** pour déposer l'Excel, puis reviens.")
    st.stop()

_, flavor_map, _ = get_paths()
df_raw = st.session_state.df_raw
window_days = st.session_state.window_days

# ---- SIDEBAR: prix moyen au choix ----
with st.sidebar:
    st.header("Paramètres pertes")
    price_hL = st.number_input(
        "Prix moyen (€/hL)",
        min_value=0.0,
        value=500.0,
        step=10.0,
        format="%.0f",
    )

st.caption(
    f"Fichier courant : **{st.session_state.get('file_name','(sans nom)')}** — "
    f"Fenêtre (B2) : **{window_days} jours** — "
    f"Prix moyen : **€{price_hL:.0f}/hL**"
)

# ---- Calculs ----
fm = load_flavor_map_from_path(flavor_map)
df_in = apply_canonical_flavor(df_raw, fm)
pertes = compute_losses_table_v48(df_in, window_days, price_hL)

colA, colB = st.columns([2, 1])
with colA:
    if isinstance(pertes, pd.DataFrame) and not pertes.empty:
        st.dataframe(pertes, use_container_width=True, hide_index=True)
    else:
        st.info("Aucune perte estimée sur 7 jours (données insuffisantes ou stock suffisant).")

with colB:
    total = float(pertes["Perte (€)"].sum()) if isinstance(pertes, pd.DataFrame) and not pertes.empty else 0.0
    kpi("Perte totale (7 j)", f"€{total:,.0f}")

-Fichier = 03_Fiche_de_ramasse.py : 
# pages/03_Fiche_de_ramasse.py
from __future__ import annotations
import os, re, datetime as dt
import unicodedata
import pandas as pd
import streamlit as st
from dateutil.tz import gettz

from common.design import apply_theme, section, kpi
# au lieu de: from common.xlsx_fill import fill_bl_enlevements_xlsx
import importlib
import common.xlsx_fill as _xlsx_fill
importlib.reload(_xlsx_fill)
from common.xlsx_fill import fill_bl_enlevements_xlsx


# ------------------------------------------------------------------
# Réglages
# ------------------------------------------------------------------
INFO_CSV_PATH = "info_FDR.csv"   # ton CSV catalogue (Code-barre, Poids, ...)
TEMPLATE_XLSX_PATH = "assets/BL_enlevements_Sofripa.xlsx"

DEST_TITLE = "SOFRIPA"
DEST_LINES = [
    "ZAC du Haut de Wissous II,",
    "Rue Hélène Boucher, 91320 Wissous",
]

# ------------------------------------------------------------------
# Utils
# ------------------------------------------------------------------
def _today_paris() -> dt.date:
    return dt.datetime.now(gettz("Europe/Paris")).date()

def _strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in s if not unicodedata.combining(ch))

def _canon(s: str) -> str:
    s = _strip_accents(str(s or "")).lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def _format_from_stock(stock_txt: str) -> str | None:
    """
    Détecte 12x33 / 6x75 / 4x75 dans un libellé de Stock.
    """
    if not stock_txt:
        return None
    s = str(stock_txt).lower().replace("×", "x").replace("\u00a0", " ")

    vol = None
    if "0.33" in s or re.search(r"33\s*c?l", s): vol = 33
    elif "0.75" in s or re.search(r"75\s*c?l", s): vol = 75

    nb = None
    m = re.search(r"(?:carton|pack)\s*de\s*(12|6|4)\b", s)
    if not m: m = re.search(r"\b(12|6|4)\b", s)
    if m: nb = int(m.group(1))

    if vol == 33 and nb == 12: return "12x33"
    if vol == 75 and nb == 6:  return "6x75"
    if vol == 75 and nb == 4:  return "4x75"
    return None

@st.cache_data(show_spinner=False)
def _load_catalog(path: str) -> pd.DataFrame:
    """
    Lit info_FDR.csv et prépare colonnes auxiliaires pour le matching.
    - normalise Poids (virgule -> point)
    - prépare Format normalisé et formes canonisées de Produit/Désignation
    """
    import pandas as pd, os, re
    if not os.path.exists(path):
        return pd.DataFrame(columns=["Produit","Format","Désignation","Code-barre","Poids"])

    df = pd.read_csv(path, encoding="utf-8")
    for c in ["Produit","Format","Désignation","Code-barre"]:
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip()

    # Poids: "7,23" -> "7.23" puis numeric
    if "Poids" in df.columns:
        df["Poids"] = (
            df["Poids"]
            .astype(str)
            .str.replace(",", ".", regex=False)
        )
        df["Poids"] = pd.to_numeric(df["Poids"], errors="coerce")

    # Format: "12x33cl" -> "12x33", "6x75cl" -> "6x75"
    df["_format_norm"] = df.get("Format","").astype(str).str.lower()
    df["_format_norm"] = (
        df["_format_norm"]
        .str.replace("cl", "", regex=False)
        .str.replace(" ", "", regex=False)
    )

    # Canon pour Produit / Désignation
    df["_canon_prod"] = df.get("Produit","").map(_canon)
    # on retire tout ce qui est entre parenthèses, puis canon
    df["_canon_des"]  = df.get("Désignation","").map(lambda s: _canon(re.sub(r"\(.*?\)", "", s)))

    return df


def _csv_lookup(catalog: pd.DataFrame, gout_canon: str, fmt_label: str) -> tuple[str, float] | None:
    """
    Retourne (référence_6_chiffres, poids_carton) en matchant :
      - format (12x33 / 6x75 / 4x75)
      - + goût canonisé (ex: 'mangue passion') contre Produit/Désignation du CSV
    """
    if catalog is None or catalog.empty or not fmt_label:
        return None

    fmt_norm = fmt_label.lower().replace("cl","").replace(" ", "")
    g_can = _canon(gout_canon)

    # filtre format d'abord
    cand = catalog[catalog["_format_norm"].str.contains(fmt_norm, na=False)]
    if cand.empty:
        return None

    # 1) match strict sur Produit canonisé
    m1 = cand[cand["_canon_prod"] == g_can]
    if m1.empty:
        # 2) sinon, on vérifie que tous les tokens du goût sont dans la désignation canonisée
        toks = [t for t in g_can.split() if t]
        def _contains_all(s):
            s2 = str(s or "")
            return all(t in s2 for t in toks)
        m1 = cand[cand["_canon_des"].map(_contains_all)]

    if m1.empty:
        # en dernier recours, on prend juste le premier du bon format
        m1 = cand

    row = m1.iloc[0]
    code = re.sub(r"\D+", "", str(row.get("Code-barre","")))
    ref6 = code[-6:] if len(code) >= 6 else code
    poids = float(row.get("Poids") or 0.0)
    return (ref6, poids) if ref6 else None


    row = cand.iloc[0]
    code = re.sub(r"\D+", "", str(row.get("Code-barre","")))
    ref6 = code[-6:] if len(code) >= 6 else code
    poids = float(row.get("Poids") or 0.0)
    return ref6, poids

# ------------------------------------------------------------------
# UI
# ------------------------------------------------------------------
apply_theme("Fiche de ramasse — Ferment Station", "🚚")
section("Fiche de ramasse", "🚚")

# Besoin de la production sauvegardée depuis la page "Production"
if "saved_production" not in st.session_state or "df_min" not in st.session_state["saved_production"]:
    st.warning("Va d’abord dans **Production** et clique **💾 Sauvegarder cette production**.")
    st.stop()

sp = st.session_state["saved_production"]
df_min_saved: pd.DataFrame = sp["df_min"].copy()
ddm_saved = dt.date.fromisoformat(sp["ddm"]) if "ddm" in sp else _today_paris()

# 1) Options dérivées de la prod sauvegardée (goût + format)
opts_rows, seen = [], set()
for _, r in df_min_saved.iterrows():
    gout = str(r.get("GoutCanon") or "").strip()
    fmt  = _format_from_stock(r.get("Stock"))
    if not (gout and fmt):
        continue
    key = (gout.lower(), fmt)
    if key in seen:
        continue
    seen.add(key)
    opts_rows.append({
        "label": f"{gout} — {fmt}",
        "gout": gout,
        "format": fmt,
        "prod_hint": str(r.get("Produit") or "").strip(),  # pour matcher le CSV
    })

if not opts_rows:
    st.error("Impossible de détecter les **formats** (12x33, 6x75, 4x75) dans la production sauvegardée.")
    st.stop()

opts_df = pd.DataFrame(opts_rows).sort_values(by="label").reset_index(drop=True)

# 2) Catalogue CSV
catalog = _load_catalog(INFO_CSV_PATH)
if catalog.empty:
    st.warning("⚠️ `info_FDR.csv` introuvable ou vide — références/poids non calculables.")

# 3) Sidebar : dates
with st.sidebar:
    st.header("Paramètres")
    date_creation = _today_paris()
    date_ramasse = st.date_input("Date de ramasse", value=date_creation)
    st.caption(f"DATE DE CRÉATION : **{date_creation.strftime('%d/%m/%Y')}**")
    st.caption(f"DDM (depuis Production) : **{ddm_saved.strftime('%d/%m/%Y')}**")

# 4) Sélection utilisateur
st.subheader("Sélection des produits")
selection_labels = st.multiselect(
    "Produits à inclure (Goût — Format)",
    options=opts_df["label"].tolist(),
    default=opts_df["label"].tolist(),
)

if not selection_labels:
    st.info("Sélectionne au moins un produit.")
    st.stop()

# 5) Prépare la table éditable (Référence + Poids issus du CSV)
meta_by_label = {}
rows = []
for lab in selection_labels:
    row_opt = opts_df.loc[opts_df["label"] == lab].iloc[0]
    gout     = row_opt["gout"]          # <-- on utilise le GOÛT canonisé
    fmt      = row_opt["format"]

    ref = ""; poids_carton = 0.0
    lk = _csv_lookup(catalog, gout, fmt)  # <-- lookup par goût + format
    if lk:
        ref, poids_carton = lk
    meta_by_label[lab] = {"_format": fmt, "_poids_carton": poids_carton, "_reference": ref}

    rows.append({
        "Référence": ref,
        "Produit (goût + format)": lab.replace(" — ", " - "),
        "DDM": ddm_saved.strftime("%d/%m/%Y"),
        "Quantité cartons": 0,
        "Quantité palettes": 0,
        "Poids palettes (kg)": 0,
    })

display_cols = ["Référence","Produit (goût + format)","DDM","Quantité cartons","Quantité palettes","Poids palettes (kg)"]
base_df = pd.DataFrame(rows, columns=display_cols)

st.caption("Renseigne **Quantité cartons** et, si besoin, **Quantité palettes**. Le **poids** se calcule automatiquement (cartons × poids/carton du CSV).")
edited = st.data_editor(
    base_df,
    key="ramasse_editor_xlsx_v1",
    use_container_width=True,
    hide_index=True,
    column_config={
        "Quantité cartons":   st.column_config.NumberColumn(min_value=0, step=1),
        "Quantité palettes":  st.column_config.NumberColumn(min_value=0, step=1),
        "Poids palettes (kg)": st.column_config.NumberColumn(disabled=True, format="%.0f"),
    },
)

# 6) Calcul poids = cartons × poids/carton
def _apply_calculs(df_disp: pd.DataFrame) -> pd.DataFrame:
    out = df_disp.copy()
    poids = []
    for _, r in out.iterrows():
        # On retrouve la clé label côté meta, avec ou sans remplacement du tiret
        lab = str(r["Produit (goût + format)"]).replace(" - ", " — ")
        meta = meta_by_label.get(lab, meta_by_label.get(str(r["Produit (goût + format)"]), {}))
        pc = float(meta.get("_poids_carton", 0.0))
        cartons = int(pd.to_numeric(r["Quantité cartons"], errors="coerce") or 0)
        poids.append(int(round(cartons * pc, 0)))
    out["Poids palettes (kg)"] = poids
    return out

df_calc = _apply_calculs(edited)

# KPIs
tot_cartons = int(pd.to_numeric(df_calc["Quantité cartons"], errors="coerce").fillna(0).sum())
tot_palettes = int(pd.to_numeric(df_calc["Quantité palettes"], errors="coerce").fillna(0).sum())
tot_poids = int(pd.to_numeric(df_calc["Poids palettes (kg)"], errors="coerce").fillna(0).sum())

c1, c2, c3 = st.columns(3)
with c1: kpi("Total cartons", f"{tot_cartons:,}".replace(",", " "))
with c2: kpi("Total palettes", f"{tot_palettes}")
with c3: kpi("Poids total (kg)", f"{tot_poids:,}".replace(",", " "))

st.dataframe(df_calc[display_cols], use_container_width=True, hide_index=True)

# 7) Téléchargement XLSX (remplissage du modèle)
st.markdown("---")
if st.button("📄 Télécharger la fiche (XLSX, modèle Sofripa)", use_container_width=True, type="primary"):
    if tot_cartons <= 0:
        st.error("Renseigne au moins une **Quantité cartons** > 0.")
    elif not os.path.exists(TEMPLATE_XLSX_PATH):
        st.error(f"Modèle Excel introuvable : `{TEMPLATE_XLSX_PATH}`")
    else:
        try:
            xlsx_bytes = fill_bl_enlevements_xlsx(
                template_path=TEMPLATE_XLSX_PATH,
                date_creation=_today_paris(),
                date_ramasse=date_ramasse,
                destinataire_title=DEST_TITLE,
                destinataire_lines=DEST_LINES,
                df_lines=df_calc[display_cols],
            )
            fname = f"BL_enlevements_{_today_paris().strftime('%Y%m%d')}.xlsx"
            st.download_button(
                "⬇️ Télécharger le XLSX",
                data=xlsx_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"Erreur lors du remplissage du modèle Excel : {e}")

- Fichier = 99_Debug.py :
# pages/99_Debug.py
import pathlib, traceback
import streamlit as st

st.set_page_config(page_title="Debug pages", page_icon="🛠️", layout="wide")
st.title("🛠️ Debug des pages Streamlit")

root = pathlib.Path(__file__).resolve().parents[1]  # racine du projet
pages = sorted((root / "pages").glob("*.py"))

bad = []
for p in pages:
    code = p.read_text(encoding="utf-8", errors="replace")
    try:
        compile(code, str(p), "exec")
        st.success(f"OK: {p.name}")
    except SyntaxError as e:
        st.error(f"SYNTAX ERROR dans {p.name} — ligne {e.lineno}, colonne {e.offset}")
        st.code("".join(traceback.format_exception_only(e)), language="text")
        # Montre la ligne incriminée
        lines = code.splitlines()
        i = max(0, (e.lineno or 1) - 1)
        snippet = "\n".join(lines[max(0, i-2): i+3])
        st.code(snippet, language="python")
        bad.append(p.name)

if not bad:
    st.info("✅ Toutes les pages compilent correctement.")
else:
    st.warning("Corrige les pages en erreur ci-dessus puis rafraîchis.")

-DOSSIER 2 = data 
-Fichier = flavor_map.csv : 
name,canonical
IGEBA Pêche, IGEBA Pêche
Infusion probiotique Mélisse - 0.0°, Infusion Mélisse 
Infusion probiotique Zest d'agrumes - 0.0°, Infusion Zest d'agrumes
Infusion probiotique menthe poivrée - 0.0°, Infusion menthe poivrée
Kéfir Gingembre, Gingembre
Kéfir Mangue Passion, Mangue Passion
Kéfir Pamplemousse, Pamplemousse
Kéfir de fruits Original, Original
Kéfir menthe citron vert, Menthe citron vert 
NIKO - Kéfir de fruits Gingembre, Gingembre
NIKO - Kéfir de fruits Mangue Passion, Mangue Passion
NIKO - Kéfir de fruits Menthe Citron-vert, Menthe citron vert 
Probiotic water Lemonbalm, Infusion Mélisse 
Probiotic water Peppermint, Infusion menthe poivrée
Water kefir Grapefruit, Pamplemousse
Water kefir Mango Passion, Mangue Passion
Water kefir Mint Lime, Menthe citron vert 

-DOSSIER 3 = core 
-Fichier = optimizer.py :
import io, re
from pathlib import Path
from typing import Optional, List, Tuple
import numpy as np
import pandas as pd

import unicodedata

def _norm_colname(s: str) -> str:
    s = str(s or "")
    s = s.strip().lower()
    # enlève accents
    s = "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))
    # remplace tout le reste par des espaces
    import re as _re
    s = _re.sub(r"[^a-z0-9]+", " ", s)
    s = _re.sub(r"\s+", " ", s).strip()
    return s

def _pick_column(df: pd.DataFrame, candidates_norm: list[str]) -> str | None:
    """
    Retourne le vrai nom de colonne du df correspondant à des candidats "normalisés".
    Amélioré : accepte 'produit 1', 'produit_2', etc. + correspondances partielles.
    """
    norm_to_real = {_norm_colname(c): c for c in df.columns}
    norms = list(norm_to_real.keys())

    # 1) match exact (priorité)
    for cand in candidates_norm:
        if cand in norm_to_real:
            return norm_to_real[cand]

    # 2) startswith sur les mots-clés importants (ex: 'produit' → 'produit 1')
    KEY_PREFIXES = ["produit", "designation", "desigation", "des", "libelle", "libelle", "product", "item", "sku"]
    for key in KEY_PREFIXES:
        for n in norms:
            if n.startswith(key):
                return norm_to_real[n]

    # 3) contains (au cas où un préfixe/ suffixe se glisse)
    for key in KEY_PREFIXES:
        for n in norms:
            if key in n:
                return norm_to_real[n]

    # 4) fuzzy (secours)
    try:
        import difflib
        match = difflib.get_close_matches(candidates_norm[0], norms, n=1, cutoff=0.85)
        if match:
            return norm_to_real[match[0]]
    except Exception:
        pass
    return None


# ======= tes constantes
ALLOWED_FORMATS = {(12, 0.33), (6, 0.75), (4, 0.75)}
ROUND_TO_CARTON = True
VOL_TOL = 0.02
EPS = 1e-9
DEFAULT_WINDOW_DAYS = 60

# ======= util accents (ton fix_text)
ACCENT_CHARS = "éèêëàâäîïôöùûüçÉÈÊËÀÂÄÎÏÔÖÙÛÜÇ"
CUSTOM_REPLACEMENTS = {
    "M�lisse": "Mélisse",
    "poivr�e": "poivrée",
    "P�che": "Pêche",
}
def _looks_better(a: str, b: str) -> bool:
    def score(s): return sum(ch in ACCENT_CHARS for ch in s)
    return score(b) > score(a)
def fix_text(s) -> str:
    if s is None: return ""
    if not isinstance(s, str): s = str(s)
    s0 = s
    try:
        s1 = s0.encode("latin1").decode("utf-8")
        if _looks_better(s0, s1): s0 = s1
    except Exception:
        pass
    if s0 in CUSTOM_REPLACEMENTS: return CUSTOM_REPLACEMENTS[s0]
    if "�" in s0: s0 = s0.replace("�", "é")
    return s0

# ======= détection en-tête & période B2 (sans Streamlit)
def detect_header_row(df_raw: pd.DataFrame) -> int:
    must = {"Produit", "Stock", "Quantité vendue", "Volume vendu (hl)", "Quantité disponible", "Volume disponible (hl)"}
    for i in range(min(10, len(df_raw))):
        if must.issubset(set(str(x).strip() for x in df_raw.iloc[i].tolist())):
            return i
    return 0

def rows_to_keep_by_fill(excel_bytes: bytes, header_idx: int) -> List[bool]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    start_row = header_idx + 2
    keep: List[bool] = []
    for r in range(start_row, ws.max_row + 1):
        is_black = False
        for cell in ws[r]:
            fill = cell.fill
            if fill and fill.fill_type:
                rgb = getattr(getattr(fill, "fgColor", None), "rgb", None) or getattr(getattr(fill, "start_color", None), "rgb", None)
                if rgb and rgb[-6:].upper() == "000000":
                    is_black = True
                    break
        keep.append(not is_black)
    return keep

def parse_days_from_b2(value) -> Optional[int]:
    try:
        if isinstance(value, (int, float)) and not pd.isna(value):
            v = int(round(float(value)));  return v if v > 0 else None
        if value is None: return None
        s = str(value).strip()
        m = re.search(r"(\d+)\s*(?:j|jour|jours)\b", s, flags=re.IGNORECASE)
        if m: return int(m.group(1)) or None
        date_pat = r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}).*?(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})"
        m2 = re.search(date_pat, s)
        if m2:
            d1 = pd.to_datetime(m2.group(1), dayfirst=True, errors="coerce")
            d2 = pd.to_datetime(m2.group(2), dayfirst=True, errors="coerce")
            if pd.notna(d1) and pd.notna(d2):
                days = int((d2 - d1).days)
                return days if days > 0 else None
        m3 = re.search(r"\b(\d{1,4})\b", s)
        if m3: 
            v = int(m3.group(1));  return v if v > 0 else None
    except Exception:
        return None
    return None

def read_input_excel_and_period_from_path(path_xlsx: str) -> Tuple[pd.DataFrame, int]:
    with open(path_xlsx, "rb") as f:
        file_bytes = f.read()
    raw = pd.read_excel(io.BytesIO(file_bytes), header=None)
    header_idx = detect_header_row(raw)
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_idx)
    keep_mask = rows_to_keep_by_fill(file_bytes, header_idx)
    if len(keep_mask) < len(df):
        keep_mask = keep_mask + [True] * (len(df) - len(keep_mask))
    df = df.iloc[[i for i, k in enumerate(keep_mask) if k]].reset_index(drop=True)
    # lecture B2
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
        b2_val = ws["B2"].value
        wd = parse_days_from_b2(b2_val)
    except Exception:
        wd = None
    return df, (wd if wd and wd > 0 else DEFAULT_WINDOW_DAYS)

# ======= flavor map
def load_flavor_map_from_path(path_csv: str) -> pd.DataFrame:
    import csv
    encodings = ["utf-8", "utf-8-sig", "cp1252", "latin1"]
    seps = [",", ";", "\t", "|"]
    if not Path(path_csv).exists():
        return pd.DataFrame(columns=["name", "canonical"])
    for enc in encodings:
        for sep in seps:
            try:
                fm = pd.read_csv(path_csv, encoding=enc, sep=sep, engine="python")
                lower = {c.lower(): c for c in fm.columns}
                if "name" in lower and "canonical" in lower:
                    fm = fm[[lower["name"], lower["canonical"]]].copy()
                    fm.columns = ["name","canonical"]
                    fm = fm.dropna()
                    fm["name"] = fm["name"].astype(str).str.strip().map(fix_text)
                    fm["canonical"] = fm["canonical"].astype(str).str.strip().map(fix_text)
                    fm = fm[(fm["name"]!="") & (fm["canonical"]!="")]
                    return fm
            except Exception:
                continue
    return pd.DataFrame(columns=["name", "canonical"])

def apply_canonical_flavor(df: pd.DataFrame, fm: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    # 1) Trouve la colonne "Produit" même si le nom diffère (Désignation, Libellé, Product, etc.)
    prod_candidates = [
    "produit", "produit 1", "produit1", "produit 2",  # tolère 'Produit 1/2'
    "designation", "désignation", "libelle", "libellé",
    "nom du produit", "product", "sku libelle", "sku libellé", "sku", "item"
    ]
    prod_candidates = [_norm_colname(x) for x in prod_candidates]
    col_prod = _pick_column(out, prod_candidates)


    if not col_prod:
        # message clair si rien n'est trouvé
        cols_list = ", ".join(map(str, out.columns))
        raise KeyError(
            "Colonne produit introuvable. "
            "Renomme la colonne en 'Produit' ou 'Désignation' (ou équivalent). "
            f"Colonnes détectées: {cols_list}"
        )

    # 2) Crée la colonne standard 'Produit'
    out["Produit"] = out[col_prod].astype(str).map(fix_text)
    out["Produit_norm"] = out["Produit"].str.strip()

    # 3) Mapping canonique (inchangé)
    if len(fm):
        fm = fm.dropna(subset=["name","canonical"]).copy()
        fm["name_norm"] = fm["name"].astype(str).map(fix_text).str.strip().str.lower()
        fm["canonical"] = fm["canonical"].astype(str).map(fix_text).str.strip()
        m_exact = dict(zip(fm["name_norm"], fm["canonical"]))
        keys = list(m_exact.keys())
        import difflib as _difflib
        def to_canonical(prod: str) -> str:
            s = str(prod).strip().lower()
            if s in m_exact: return m_exact[s]
            try:
                close = _difflib.get_close_matches(s, keys, n=1, cutoff=0.92)
                if close: return m_exact[close[0]]
            except Exception:
                pass
            return str(prod).strip()
        out["GoutCanon"] = out["Produit_norm"].map(to_canonical)
    else:
        out["GoutCanon"] = out["Produit_norm"]

    out["GoutCanon"] = out["GoutCanon"].astype(str).map(fix_text).str.strip()
    return out


# ======= parsing formats/stock & filtres
def parse_stock(text: str):
    if pd.isna(text): return np.nan, np.nan
    s = str(text)
    nb = None
    for pat in [r"(?:Carton|Caisse|Colis)\s+de\s*(\d+)", r"(\d+)\s*[x×]\s*Bouteilles?", r"(\d+)\s*[x×]", r"(\d+)\s+Bouteilles?"]:
        m = re.search(pat, s, flags=re.IGNORECASE)
        if m:
            try: nb = int(m.group(1)); break
            except: pass
    vol_l = None
    m_l = re.findall(r"(\d+(?:[.,]\d+)?)\s*[lL]", s)
    if m_l: vol_l = float(m_l[-1].replace(",", "."))
    else:
        m_cl = re.findall(r"(\d+(?:[.,]\d+)?)\s*c[lL]", s)
        if m_cl: vol_l = float(m_cl[-1].replace(",", ".")) / 100.0
    if nb is None or vol_l is None:
        m_combo = re.search(r"(\d+)\s*[x×]\s*(\d+(?:[.,]\d+)?)+\s*([lc]l?)", s, flags=re.IGNORECASE)
        if m_combo:
            try:
                nb2 = int(m_combo.group(1)); val = float(m_combo.group(2).replace(",", "."))
                unit = m_combo.group(3).lower(); vol2 = val if unit.startswith("l") else val/100.0
                if nb is None: nb = nb2
                if vol_l is None: vol_l = vol2
            except: pass
    if (nb is None or np.isnan(nb)) and vol_l is not None and abs(vol_l - 0.75) <= VOL_TOL:
        if re.search(r"(?:\b4\s*[x×]\b|Carton\s+de\s*4\b|4\s+Bouteilles?)", s, flags=re.IGNORECASE):
            nb = 4
    return (float(nb) if nb is not None else np.nan, float(vol_l) if vol_l is not None else np.nan)

def safe_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")

def is_allowed_format(nb_bottles, vol_l, stock_txt: str) -> bool:
    if pd.isna(nb_bottles) or pd.isna(vol_l):
        if re.search(r"(?:\b4\s*[x×]\s*75\s*c?l\b|\b4\s+Bouteilles?\b.*75\s*c?l)", stock_txt, flags=re.IGNORECASE):
            nb_bottles = 4; vol_l = 0.75
        else:
            return False
    nb_bottles = int(nb_bottles); vol_l = float(vol_l)
    for nb_ok, vol_ok in ALLOWED_FORMATS:
        if nb_bottles == nb_ok and abs(vol_l - vol_ok) <= VOL_TOL:
            return True
    return False

BLOCKED_LABELS_EXACT = {"Autres (coffrets, goodies...)"}
BLOCKED_LABELS_LOWER = {"nan", "none", ""}

def sanitize_gouts(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["GoutCanon"] = out["GoutCanon"].astype(str).str.strip()
    mask = ~out["GoutCanon"].str.lower().isin(BLOCKED_LABELS_LOWER)
    mask &= ~out["GoutCanon"].isin(BLOCKED_LABELS_EXACT)
    return out.loc[mask].reset_index(drop=True)

# ======= tes calculs (inchangés)
def compute_plan(df_in, window_days, volume_cible, nb_gouts, repartir_pro_rv, manual_keep, exclude_list):
    required = ["Produit", "GoutCanon", "Stock", "Quantité vendue", "Volume vendu (hl)", "Quantité disponible", "Volume disponible (hl)"]
    miss = [c for c in required if c not in df_in.columns]
    if miss: raise ValueError(f"Colonnes manquantes: {miss}")

    # --- helper catégorie (NEW) ---
    def _category(g: str) -> str:
        s = str(g or "").strip().lower()
        # Tout ce qui contient 'infusion' -> infusion ; sinon on considère 'kéfir'
        # (tes canoniques sont du type 'Infusion Mélisse', 'Mangue Passion', 'Gingembre', etc.)
        return "infusion" if "infusion" in s else "kefir"

    note_msg = ""  # NEW: message d’ajustement à renvoyer à l’UI

    df = df_in[required].copy()
    for c in ["Quantité vendue", "Volume vendu (hl)", "Quantité disponible", "Volume disponible (hl)"]:
        df[c] = safe_num(df[c])

    parsed = df["Stock"].apply(parse_stock)
    df[["Bouteilles/carton", "Volume bouteille (L)"]] = pd.DataFrame(parsed.tolist(), index=df.index)
    mask_allowed = df.apply(lambda r: is_allowed_format(r["Bouteilles/carton"], r["Volume bouteille (L)"], str(r["Stock"])), axis=1)
    df = df.loc[mask_allowed].reset_index(drop=True)

    df["Volume/carton (hL)"] = (df["Bouteilles/carton"] * df["Volume bouteille (L)"]) / 100.0
    df = df.dropna(subset=["GoutCanon", "Volume/carton (hL)", "Volume vendu (hl)", "Volume disponible (hl)"]).reset_index(drop=True)

    df_all_formats = df.copy()

    if exclude_list:
        ex = {s.strip() for s in exclude_list}
        df = df[~df["GoutCanon"].astype(str).str.strip().isin(ex)]

    if manual_keep:
        keep = {g.strip() for g in manual_keep}
        df = df[df["GoutCanon"].astype(str).str.strip().isin(keep)]

    agg = df.groupby("GoutCanon").agg(
        ventes_hl=("Volume vendu (hl)", "sum"),
        stock_hl=("Volume disponible (hl)", "sum")
    )
    agg["vitesse_j"] = agg["ventes_hl"] / max(float(window_days), 1.0)
    agg["jours_autonomie"] = np.where(agg["vitesse_j"] > 0, agg["stock_hl"] / agg["vitesse_j"], np.inf)
    agg["score_urgence"] = agg["vitesse_j"] / (agg["jours_autonomie"] + EPS)
    # tri par urgence décroissante
    agg = agg.sort_values(by=["score_urgence", "jours_autonomie", "ventes_hl"], ascending=[False, True, False])

    # --- Sélection initiale (comme avant) ---
    if not manual_keep:
        gouts_cibles = agg.index.tolist()[:nb_gouts]
    else:
        gouts_cibles = sorted(set(df["GoutCanon"]))
        if len(gouts_cibles) > nb_gouts:
            order = [g for g in agg.index if g in gouts_cibles]
            gouts_cibles = order[:nb_gouts]

    # --- Ajustement "pas de mix infusion+kefir si nb_gouts=2" (NEW) ---
    if nb_gouts == 2 and len(gouts_cibles) == 2:
        cat_set = { _category(g) for g in gouts_cibles }
        if len(cat_set) > 1:
            # On cherche la meilleure paire de même catégorie en respectant l'ordre d'urgence d'agg
            # 1) Catégorie du plus urgent (1er de agg parmi les deux)
            ordered = [g for g in agg.index if g in set(gouts_cibles)]
            first = ordered[0] if ordered else gouts_cibles[0]
            target_cat = _category(first)

            # 2) Liste candidates de la catégorie choisie (dans l'ordre agg)
            same_cat_all = [g for g in agg.index if _category(g) == target_cat]
            if len(same_cat_all) >= 2:
                new_pair = same_cat_all[:2]
                if set(new_pair) != set(gouts_cibles):
                    note_msg = (
                        "⚠️ Contrainte appliquée : pas de co-production **Infusion + Kéfir**. "
                        f"Sélection ajustée → deux recettes **{ 'Infusion' if target_cat=='infusion' else 'Kéfir' }** "
                        f"({new_pair[0]} ; {new_pair[1]})."
                    )
                gouts_cibles = new_pair
            else:
                # Pas 2 goûts dans la même catégorie que le plus urgent → on tente l'autre catégorie
                other_cat = "kefir" if target_cat == "infusion" else "infusion"
                other_all = [g for g in agg.index if _category(g) == other_cat]
                if len(other_all) >= 2:
                    gouts_cibles = other_all[:2]
                    note_msg = (
                        "⚠️ Contrainte appliquée : pas de co-production **Infusion + Kéfir**. "
                        "La catégorie initiale ne contenait pas 2 goûts disponibles ; "
                        f"sélection basculée sur deux recettes **{ 'Infusion' if other_cat=='infusion' else 'Kéfir' }** "
                        f"({gouts_cibles[0]} ; {gouts_cibles[1]})."
                    )
                else:
                    # Impossible de satisfaire strictement la contrainte (ex: 1 seul goût total)
                    note_msg = (
                        "⚠️ Contrainte non pleinement satisfaisable : moins de deux goûts disponibles dans une même recette. "
                        "Vérifie les filtres/forçages ou ajoute des produits."
                    )

    df_selected = df[df["GoutCanon"].isin(gouts_cibles)].copy()
    if len(gouts_cibles) == 0:
        raise ValueError("Aucun goût sélectionné.")

    df_calc = df_selected.copy()
    if nb_gouts == 1:
        df_calc["Somme ventes (hL) par goût"] = df_calc.groupby("GoutCanon")["Volume vendu (hl)"].transform("sum")
        if repartir_pro_rv:
            df_calc["r_i"] = np.where(df_calc["Somme ventes (hL) par goût"] > 0,
                                      df_calc["Volume vendu (hl)"] / df_calc["Somme ventes (hL) par goût"], 0.0)
        else:
            df_calc["r_i"] = 1.0 / df_calc.groupby("GoutCanon")["GoutCanon"].transform("count")

        df_calc["G_i (hL)"] = df_calc["Volume disponible (hl)"]
        df_calc["G_total (hL) par goût"] = df_calc.groupby("GoutCanon")["G_i (hL)"].transform("sum")
        df_calc["Y_total (hL) par goût"] = df_calc["G_total (hL) par goût"] + float(volume_cible)
        df_calc["X_th (hL)"] = df_calc["r_i"] * df_calc["Y_total (hL) par goût"] - df_calc["G_i (hL)"]

        df_calc["X_adj (hL)"] = 0.0
        for _, grp in df_calc.groupby("GoutCanon"):
            x = grp["X_th (hL)"].to_numpy(float)
            r = grp["r_i"].to_numpy(float)
        
            # clamp à 0 (on ne "déproduit" pas)
            x = np.maximum(x, 0.0)
        
            # somme vs cible
            V = float(volume_cible)
            sum_x = x.sum()
            deficit = V - sum_x
        
            if deficit > 1e-9:
                # on complète au prorata r
                r = np.where(r > 0, r, 0); s = r.sum()
                x = x + (deficit * (r / s) if s > 0 else deficit / max(len(x), 1))
            elif deficit < -1e-9:
                # on a dépassé V suite au clamp -> réduction proportionnelle
                x = x * (V / max(sum_x, 1e-12))
        
            x = np.where(x < 1e-9, 0.0, x)
            df_calc.loc[grp.index, "X_adj (hL)"] = x

        cap_resume = f"{volume_cible:.2f} hL par goût"
    else:
        # --------- PARTAGE DE LA CIBLE ENTRE LES GOÛTS (phase 1) ----------
        V = float(volume_cible)
        df_calc["G_i (hL)"] = df_calc["Volume disponible (hl)"]

        # poids par goût (ventes si dispo, sinon égalitaire)
        ventes_par_gout = df_calc.groupby("GoutCanon")["Volume vendu (hl)"].sum()
        n_gouts = max(len(ventes_par_gout), 1)
        
        # ⚖️ Partage entre goûts :
        # - si TOUS les goûts ont des ventes > 0 et que "prorata" est coché → prorata des ventes
        # - sinon → partage égalitaire (assure un volume non nul pour chaque goût sélectionné)
        if repartir_pro_rv and float(ventes_par_gout.sum()) > 0 and (ventes_par_gout > 0).all():
            w_gout = ventes_par_gout / ventes_par_gout.sum()
        else:
            w_gout = pd.Series(1.0 / n_gouts, index=ventes_par_gout.index)


        # --------- POIDS À L’INTÉRIEUR DE CHAQUE GOÛT (phase 2) ----------
        def _weights_inside(grp_df: pd.DataFrame) -> pd.Series:
            if repartir_pro_rv:
                s = float(grp_df["Volume vendu (hl)"].sum())
                if s > 0:
                    return grp_df["Volume vendu (hl)"] / s
            # égalitaire si pas de ventes
            n = max(len(grp_df), 1)
            return pd.Series([1.0 / n] * n, index=grp_df.index)

        df_calc["w_in"] = (
            df_calc.groupby("GoutCanon", group_keys=False)
                   .apply(_weights_inside)
        )

        # --------- ALLOCATION PAR GOÛT (respect strict de la cible V) -----
        df_calc["X_adj (hL)"] = 0.0
        for g, grp in df_calc.groupby("GoutCanon"):
            idx = grp.index
            Vg = V * float(w_gout.get(g, 0.0))  # cible pour ce goût
            Gg = float(grp["G_i (hL)"].sum())   # stock total de ce goût

            w_in = grp["w_in"].to_numpy(float)
            Gi   = grp["G_i (hL)"].to_numpy(float)

            # besoin théorique ligne: w_in * (Gg + Vg) - Gi
            x = w_in * (Gg + Vg) - Gi
            x = np.maximum(x, 0.0)

            sum_x = x.sum()
            if sum_x < Vg - 1e-9:
                # on complète au prorata des poids internes (même pour les lignes à 0)
                s = w_in.sum()
                add = Vg - sum_x
                x = x + (add * (w_in / s) if s > 0 else add / max(len(x), 1))
            elif sum_x > Vg + 1e-9:
                # on réduit proportionnellement pour revenir à la cible
                x = x * (Vg / sum_x)

            x = np.where(x < 1e-9, 0.0, x)
            df_calc.loc[idx, "X_adj (hL)"] = x

        cap_resume = f"{volume_cible:.2f} hL au total (2 goûts)"


    df_calc["Cartons à produire (exact)"] = df_calc["X_adj (hL)"] / df_calc["Volume/carton (hL)"]
    if ROUND_TO_CARTON:
        df_calc["Cartons à produire (arrondi)"] = np.floor(df_calc["Cartons à produire (exact)"] + 0.5).astype("Int64")
        df_calc["Volume produit arrondi (hL)"] = df_calc["Cartons à produire (arrondi)"] * df_calc["Volume/carton (hL)"]

    df_calc["Bouteilles à produire (exact)"] = df_calc["Cartons à produire (exact)"] * df_calc["Bouteilles/carton"]
    if ROUND_TO_CARTON:
        df_calc["Bouteilles à produire (arrondi)"] = (
            df_calc["Cartons à produire (arrondi)"] * df_calc["Bouteilles/carton"]
        ).astype("Int64")

    df_min = df_calc[[
        "GoutCanon", "Produit", "Stock",
        "Cartons à produire (arrondi)",
        "Bouteilles à produire (arrondi)",
        "Volume produit arrondi (hL)"
    ]].sort_values(["GoutCanon", "Produit", "Stock"]).reset_index(drop=True)

    agg_full = df.groupby("GoutCanon").agg(
        ventes_hl=("Volume vendu (hl)", "sum"),
        stock_hl=("Volume disponible (hl)", "sum")
    )
    agg_full["vitesse_j"] = agg_full["ventes_hl"] / max(float(window_days), 1.0)
    agg_full["jours_autonomie"] = np.where(agg_full["vitesse_j"] > 0, agg_full["stock_hl"] / agg_full["vitesse_j"], np.inf)
    agg_full["score_urgence"] = agg_full["vitesse_j"] / (agg_full["jours_autonomie"] + EPS)
    sel_gouts = sorted(set(df_calc["GoutCanon"]))
    synth_sel = agg_full.loc[sel_gouts][["ventes_hl", "stock_hl", "vitesse_j", "jours_autonomie", "score_urgence"]].copy()
    synth_sel = synth_sel.rename(columns={
        "ventes_hl": "Ventes 2 mois (hL)",
        "stock_hl": "Stock (hL)",
        "vitesse_j": "Vitesse (hL/j)",
        "jours_autonomie": "Autonomie (jours)",
        "score_urgence": "Score urgence"
    })
    # NEW: on renvoie note_msg en 7e sortie
    return df_min, cap_resume, sel_gouts, synth_sel, df_calc, df, note_msg

def compute_losses_table_v48(df_in_all: pd.DataFrame, window_days: float, price_hL: float) -> pd.DataFrame:
    out_cols = ["Goût", "Demande 7 j (hL)", "Stock (hL)", "Manque sur 7 j (hL)", "Prix moyen (€/hL)", "Perte (€)"]
    if df_in_all is None or not isinstance(df_in_all, pd.DataFrame) or df_in_all.empty:
        return pd.DataFrame(columns=out_cols)
    df = df_in_all.copy()
    if "GoutCanon" not in df.columns:
        return pd.DataFrame(columns=out_cols)
    for c in ["Quantité vendue", "Volume vendu (hl)", "Quantité disponible", "Volume disponible (hl)"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    df["GoutCanon"] = df["GoutCanon"].astype(str).str.strip()
    bad_lower = {"nan", "none", ""}
    df = df[~df["GoutCanon"].str.lower().isin(bad_lower)]
    df = df[df["GoutCanon"] != "Autres (coffrets, goodies...)"]
    if df.empty:
        return pd.DataFrame(columns=out_cols)
    jours = max(float(window_days), 1.0)
    agg = df.groupby("GoutCanon", as_index=False).agg(
        ventes_hL=("Volume vendu (hl)", "sum"),
        stock_hL=("Volume disponible (hl)", "sum"),
    )
    if agg.empty:
        return pd.DataFrame(columns=out_cols)
    agg["vitesse_hL_j"] = agg["ventes_hL"] / jours
    agg["Demande 7 j (hL)"] = 7.0 * agg["vitesse_hL_j"]
    agg["Stock (hL)"] = agg["stock_hL"]
    agg["Manque sur 7 j (hL)"] = np.clip(agg["Demande 7 j (hL)"] - agg["Stock (hL)"], a_min=0.0, a_max=None)
    agg["Prix moyen (€/hL)"] = float(price_hL)
    agg["Perte (€)"] = (agg["Manque sur 7 j (hL)"] * agg["Prix moyen (€/hL)"]).round(0)
    pertes = agg.rename(columns={"GoutCanon": "Goût"})[
        ["Goût", "Demande 7 j (hL)", "Stock (hL)", "Manque sur 7 j (hL)", "Prix moyen (€/hL)", "Perte (€)"]
    ]
    pertes["Goût"] = pertes["Goût"].map(fix_text)
    pertes["Demande 7 j (hL)"] = pertes["Demande 7 j (hL)"].round(2)
    pertes["Stock (hL)"] = pertes["Stock (hL)"].round(2)
    pertes["Manque sur 7 j (hL)"] = pertes["Manque sur 7 j (hL)"].round(2)
    pertes["Prix moyen (€/hL)"] = pertes["Prix moyen (€/hL)"].round(0)
    return pertes.sort_values("Perte (€)", ascending=False).reset_index(drop=True)

# --- LECTURE EXCEL depuis un UPLOAD Streamlit (sans rien changer ailleurs) ---

def read_input_excel_and_period_from_bytes(file_bytes: bytes):
    """Même logique que _from_path mais pour des bytes (uploader Streamlit)."""
    import io, openpyxl
    raw = pd.read_excel(io.BytesIO(file_bytes), header=None)
    header_idx = detect_header_row(raw)
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_idx)
    keep_mask = rows_to_keep_by_fill(file_bytes, header_idx)
    if len(keep_mask) < len(df):
        keep_mask = keep_mask + [True] * (len(df) - len(keep_mask))
    df = df.iloc[[i for i, k in enumerate(keep_mask) if k]].reset_index(drop=True)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
        b2_val = ws["B2"].value
        wd = parse_days_from_b2(b2_val)
    except Exception:
        wd = None
    return df, (wd if wd and wd > 0 else DEFAULT_WINDOW_DAYS)

def read_input_excel_and_period_from_upload(uploaded_file):
    """Wrapper pratique pour st.file_uploader (obj upload Streamlit)."""
    file_bytes = uploaded_file.read()
    return read_input_excel_and_period_from_bytes(file_bytes)

- DOSSIER 4 = common 
-Fichier = xlsx_fill.py : 
# common/xlsx_fill.py
from __future__ import annotations

import io
import os
import re
import unicodedata
from datetime import date, datetime
from typing import Optional, Dict, List, Tuple

from dateutil.relativedelta import relativedelta
import pandas as pd
import openpyxl
from openpyxl.utils import coordinate_to_tuple, get_column_letter

# ======================================================================
#                         Utilitaires généraux
# ======================================================================

VOL_TOL = 0.02

def _is_close(a: float, b: float, tol: float = VOL_TOL) -> bool:
    try:
        return abs(float(a) - float(b)) <= tol
    except Exception:
        return False

# ----------- parse format depuis la colonne "Stock" (df_min) -----------
def _parse_format_from_stock(stock: str):
    s = str(stock or "")
    m_nb = re.search(r'(Carton|Pack)\s+de\s+(\d+)\s+Bouteilles?', s, flags=re.I)
    nb = int(m_nb.group(2)) if m_nb else None
    m_l = re.search(r'(\d+(?:[.,]\d+)?)\s*[lL]\b', s)
    vol = float(m_l.group(1).replace(",", ".")) if m_l else None
    if vol is None:
        m_cl = re.search(r'(\d+(?:[.,]\d+)?)\s*c[lL]\b', s)
        vol = float(m_cl.group(1).replace(",", "."))/100.0 if m_cl else None
    return nb, vol

# ----------- Agrégat STRICT depuis df_min (tableau affiché) -----------
def _agg_from_dfmin(df_min: pd.DataFrame, gout: str) -> Dict[str, Dict[str, int]]:
    out = {
        "33_fr":  {"cartons": 0, "bouteilles": 0},
        "33_niko":{"cartons": 0, "bouteilles": 0},
        "75x6":   {"cartons": 0, "bouteilles": 0},
        "75x4":   {"cartons": 0, "bouteilles": 0},
    }
    if df_min is None or not isinstance(df_min, pd.DataFrame) or df_min.empty:
        return out
    req = {"Produit","Stock","GoutCanon","Cartons à produire (arrondi)","Bouteilles à produire (arrondi)"}
    if any(c not in df_min.columns for c in req):
        return out

    df = df_min.copy()
    df = df[df["GoutCanon"].astype(str).str.strip() == str(gout).strip()]
    if df.empty:
        return out

    for _, r in df.iterrows():
        nb, vol = _parse_format_from_stock(r["Stock"])
        if nb is None or vol is None:
            continue
        ct = int(pd.to_numeric(r["Cartons à produire (arrondi)"], errors="coerce") or 0)
        bt = int(pd.to_numeric(r["Bouteilles à produire (arrondi)"], errors="coerce") or 0)
        prod_up = str(r["Produit"]).upper()

        if nb == 12 and _is_close(vol, 0.33):
            key = "33_niko" if "NIKO" in prod_up else "33_fr"
        elif nb == 6 and _is_close(vol, 0.75):
            key = "75x6"
        elif nb == 4 and _is_close(vol, 0.75):
            key = "75x4"
        else:
            continue

        out[key]["cartons"]    += ct
        out[key]["bouteilles"] += bt

    return out

# ----------- Helper écriture tolérante aux fusions -----------
def _set(ws, addr: str, value, number_format: str | None = None):
    row, col = coordinate_to_tuple(addr)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            row, col = rng.min_row, rng.min_col
            break
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if number_format:
        cell.number_format = number_format
    return f"{get_column_letter(col)}{row}"

# ----------- Détection auto des blocs Quantité (paire du BAS) -----------
def _norm(s) -> str:
    return str(s).strip().lower()

def _locate_quantity_blocks(ws) -> Dict[str, Dict[str, int]]:
    """
    Le modèle contient 2 paires de blocs (haut = résumé, bas = zone d'entrée).
    On retourne **la paire du BAS** pour la saisie.
    """
    labels = {"france", "niko", "x6", "x4"}
    row_hits: Dict[int, Dict[str, int]] = {}

    for r in ws.iter_rows(values_only=False):
        for c in r:
            v = c.value
            if isinstance(v, str):
                nv = _norm(v)
                if nv in labels:
                    row_hits.setdefault(c.row, {})[nv] = c.column

    candidates = [(row, cols) for row, cols in row_hits.items() if len(cols) >= 3]
    if len(candidates) < 2:
        raise KeyError("En-têtes 'France/NIKO/X6/X4' introuvables (paire du bas non détectée).")

    # On prend les 2 lignes les plus basses (bas de page)
    candidates.sort(key=lambda x: x[0])
    bottom_pair = candidates[-2:]

    def _avg_col(cols: Dict[str, int]) -> float:
        return sum(cols.values()) / len(cols)

    # gauche / droite
    bottom_pair.sort(key=lambda x: _avg_col(x[1]))
    (left_row, left_cols), (right_row, right_cols) = bottom_pair

    def _fill_missing(cols: Dict[str, int]) -> Dict[str, int]:
        out = cols.copy()
        for k in ["france", "niko", "x6", "x4"]:
            out.setdefault(k, next(iter(out.values())))
        return out

    left_cols  = _fill_missing(left_cols)
    right_cols = _fill_missing(right_cols)

    return {
        "left":  {"header_row": left_row,  "bouteilles_row": left_row + 1, "cartons_row": left_row + 2, **left_cols},
        "right": {"header_row": right_row, "bouteilles_row": right_row + 1, "cartons_row": right_row + 2, **right_cols},
    }

def _addr(col: int, row: int) -> str:
    return f"{get_column_letter(col)}{row}"

# ======================================================================
#     Fiche de prod 7000L (existante dans ton repo) — inchangée
# ======================================================================

def fill_fiche_7000L_xlsx(
    template_path: str,
    semaine_du: date,
    ddm: date,
    gout1: str,
    gout2: Optional[str],
    df_calc,
    sheet_name: str | None = None,
    df_min=None,
) -> bytes:
    wb = openpyxl.load_workbook(template_path, data_only=False, keep_vba=False)

    targets = [sheet_name] if sheet_name else ["Fiche de production 7000 L", "Fiche de production 7000L"]
    ws = None
    for nm in targets:
        if nm and nm in wb.sheetnames:
            ws = wb[nm]
            break
    if ws is None:
        raise KeyError(f"Feuille cible introuvable. Feuilles présentes : {wb.sheetnames}")

    # En-têtes
    _set(ws, "D8", gout1 or "")
    _set(ws, "T8", gout2 or "")
    _set(ws, "D10", ddm, number_format="DD/MM/YYYY")
    _set(ws, "O10", ddm.strftime("%d%m%Y"))
    ferment_date = ddm - relativedelta(years=1)
    _set(ws, "A20", ferment_date, number_format="DD/MM/YYYY")

    # Localisation des blocs
    blocks = _locate_quantity_blocks(ws)
    L = blocks["left"];  R = blocks["right"]

    P1 = {
        "33_fr":  {"b": _addr(L["france"], L["bouteilles_row"]), "c": _addr(L["france"], L["cartons_row"])},
        "33_niko":{"b": _addr(L["niko"],   L["bouteilles_row"]), "c": _addr(L["niko"],   L["cartons_row"])},
        "75x6":   {"b": _addr(L["x6"],     L["bouteilles_row"]), "c": _addr(L["x6"],     L["cartons_row"])},
        "75x4":   {"b": _addr(L["x4"],     L["bouteilles_row"]), "c": _addr(L["x4"],     L["cartons_row"])},
    }
    P2 = {
        "33_fr":  {"b": _addr(R["france"], R["bouteilles_row"]), "c": _addr(R["france"], R["cartons_row"])},
        "33_niko":{"b": _addr(R["niko"],   R["bouteilles_row"]), "c": _addr(R["niko"],   R["cartons_row"])},
        "75x6":   {"b": _addr(R["x6"],     R["bouteilles_row"]), "c": _addr(R["x6"],     R["cartons_row"])},
        "75x4":   {"b": _addr(R["x4"],     R["bouteilles_row"]), "c": _addr(R["x4"],     R["cartons_row"])},
    }

    # --- Agrégats : df_min uniquement (copie EXACTE du tableau affiché)
    agg1 = _agg_from_dfmin(df_min, gout1)
    agg2 = _agg_from_dfmin(df_min, gout2) if gout2 else None

    # N'écrit rien si 0 → on laisse les pointillés du modèle
    def _write_if_pos(addr: str, val):
        v = int(pd.to_numeric(val, errors="coerce") or 0)
        if v > 0:
            _set(ws, addr, v)

    # Gauche (Produit 1)
    for k, dest in P1.items():
        _write_if_pos(dest["b"], agg1[k]["bouteilles"])
        _write_if_pos(dest["c"], agg1[k]["cartons"])

    # Droite (Produit 2) si présent (sinon on ne touche pas aux pointillés)
    if agg2 is not None:
        for k, dest in P2.items():
            _write_if_pos(dest["b"], agg2[k]["bouteilles"])
            _write_if_pos(dest["c"], agg2[k]["cartons"])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ======================================================================
#                   Remplissage BL enlèvements Sofripa
# ======================================================================

# --- Helpers dédiés au modèle BL Sofripa ---

def _iter_cells(ws):
    for r in ws.iter_rows(values_only=False):
        for c in r:
            yield c

def _find_cell_by_regex(ws, pattern: str) -> Tuple[int, int] | Tuple[None, None]:
    rx = re.compile(pattern, flags=re.I)
    for cell in _iter_cells(ws):
        v = cell.value
        if isinstance(v, str) and rx.search(v):
            return cell.row, cell.column
    return None, None

def _write_right_of(ws, row: int, col: int, value):
    ws.cell(row=row, column=col + 1).value = value

def _normalize_header_text(s: str) -> str:
    s = str(s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace("’", "'")
    for ch in ["(", ")", ":", ";", ","]:
        s = s.replace(ch, " ")
    s = " ".join(s.split())
    return s

def _find_table_headers(ws, targets: List[str]) -> Tuple[int | None, Dict[str, int]]:
    """
    Essaie de trouver une ligne qui ressemble à des en-têtes du tableau principal.
    Retourne (row_index, mapping_nom->col_index_1based)
    """
    norm_targets = [_normalize_header_text(t) for t in targets]

    # on parcourt les premières ~50 lignes pour trouver une majorité de correspondances
    best_row = None
    best_map: Dict[str, int] = {}
    max_hits = 0

    max_rows = min(ws.max_row, 80)
    max_cols = min(ws.max_column, 50)

    for r in range(1, max_rows + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_cols + 1)]
        row_norm = [_normalize_header_text(x) for x in row_vals]

        colmap: Dict[str, int] = {}
        hits = 0
        for t_norm, t_orig in zip(norm_targets, targets):
            found = False
            for j, hv in enumerate(row_norm, start=1):
                if hv == t_norm:
                    colmap[t_orig] = j
                    hits += 1
                    found = True
                    break
            if not found:
                # essais souples (contains)
                for j, hv in enumerate(row_norm, start=1):
                    if t_norm in hv and len(t_norm) >= 4:
                        colmap[t_orig] = j
                        hits += 1
                        found = True
                        break

        if hits > max_hits:
            max_hits = hits
            best_row = r
            best_map = colmap

        if hits >= len(targets) - 1:  # quasi toutes
            break

    return best_row, best_map


def fill_bl_enlevements_xlsx(
    template_path: str,
    date_creation: date,
    date_ramasse: date,
    destinataire_title: str,
    destinataire_lines: List[str],
    df_lines: pd.DataFrame,   # colonnes attendues (ordre libre) cf. ci-dessous
) -> bytes:
    """
    Remplit le modèle XLSX 'LOG_EN_001_01 BL enlèvements Sofripa-2.xlsx'.

    df_lines doit contenir les colonnes (noms exacts ou équivalents) :
      - 'Référence'
      - 'Produit (goût + format)' ou 'Produit'
      - 'DDM'
      - 'Quantité cartons'
      - 'Quantité palettes'
      - 'Poids palettes (kg)'
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Modèle Excel introuvable: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # ----- 1) Dates -----
    r, c = _find_cell_by_regex(ws, r"date\s+de\s+cr[eé]ation")
    if r and c:
        _write_right_of(ws, r, c, date_creation.strftime("%d/%m/%Y"))

    r, c = _find_cell_by_regex(ws, r"date\s+de\s+rammasse|date\s+de\s+ramasse")
    if r and c:
        _write_right_of(ws, r, c, date_ramasse.strftime("%d/%m/%Y"))

    # ----- 2) Destinataire -----
    r, c = _find_cell_by_regex(ws, r"destinataire")
    if r and c:
        _write_right_of(ws, r, c, destinataire_title)
        for i, line in enumerate(destinataire_lines[:3], start=1):
            ws.cell(row=r + i, column=c + 1).value = line

    # ----- 3) En-têtes du tableau (tolérant) -----
    hdr_row, _ = _find_table_headers(ws, [
        "Référence", "Produit", "DDM",
        "Quantité cartons", "Quantité palettes", "Poids palettes (kg)"
    ])
    if not hdr_row:
        raise KeyError("Ligne d’en-têtes du tableau introuvable dans le modèle Excel.")

    header_vals = [ws.cell(row=hdr_row, column=j).value for j in range(1, ws.max_column + 1)]

    def _match_header(target: str, contains: bool=False) -> int | None:
        t = _normalize_header_text(target)
        for j, v in enumerate(header_vals, start=1):
            h = _normalize_header_text(v)
            if (contains and t in h) or (not contains and t == h):
                return j
        return None

    c_ref   = _match_header("référence") or _match_header("reference")
    c_prod  = (_match_header("produit")
               or _match_header("produit (gout + format)", contains=True)
               or _match_header("produit gout format", contains=True))
    c_ddm   = _match_header("ddm") or _match_header("date de durabilite", contains=True)
    c_qc    = _match_header("quantité cartons") or _match_header("quantite cartons")
    c_qp    = _match_header("quantité palettes") or _match_header("quantite palettes")
    c_poids = (_match_header("poids palettes (kg)")
               or _match_header("poids palettes")
               or _match_header("poids (kg)"))

    # fallback : Produit entre Réf et DDM
    if c_prod is None and c_ref is not None and c_ddm is not None and c_ddm > c_ref:
        if (c_ddm - c_ref) >= 2:
            c_prod = c_ref + 1

    need = {
        "Référence": c_ref,
        "Produit": c_prod,
        "DDM": c_ddm,
        "Quantité cartons": c_qc,
        "Quantité palettes": c_qp,
        "Poids palettes (kg)": c_poids,
    }
    if any(v is None for v in need.values()):
        raise ValueError(f"Colonnes incomplètes dans le modèle Excel: {need}")

    # ----- 4) Normalisation DF d'entrée -----
    df = df_lines.copy()
    # alias Produit
    if "Produit" not in df.columns and "Produit (goût + format)" in df.columns:
        df = df.rename(columns={"Produit (goût + format)": "Produit"})
    # DDM → texte jj/mm/aaaa
    def _to_ddm_val(x):
        if isinstance(x, (date, )):
            return x.strftime("%d/%m/%Y")
        s = str(x or "").strip()
        if not s:
            return ""
        # supports "yyyy-mm-dd" or "dd/mm/yyyy"
        try:
            if "-" in s and len(s.split("-")[0]) == 4:
                return datetime.strptime(s, "%Y-%m-%d").strftime("%d/%m/%Y")
            return datetime.strptime(s, "%d/%m/%Y").strftime("%d/%m/%Y")
        except Exception:
            return s

    # ----- 5) Écriture des lignes -----
    row = hdr_row + 1

    def _as_int(v) -> int:
        try:
            f = float(v)
            return int(round(f))
        except Exception:
            return 0

    for _, r in df.iterrows():
        ws.cell(row=row, column=c_ref).value   = str(r.get("Référence", ""))
        ws.cell(row=row, column=c_prod).value  = str(r.get("Produit", ""))
        ws.cell(row=row, column=c_ddm).value   = _to_ddm_val(r.get("DDM", ""))

        ws.cell(row=row, column=c_qc).value    = _as_int(r.get("Quantité cartons", 0))
        ws.cell(row=row, column=c_qp).value    = _as_int(r.get("Quantité palettes", 0))
        ws.cell(row=row, column=c_poids).value = _as_int(r.get("Poids palettes (kg)", 0))
        row += 1

    # ----- 6) Sauvegarde -----
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

- Fichier = design.py : 
import re, unicodedata, os
from io import BytesIO
from PIL import Image
import streamlit as st

COLORS = {
    "bg": "#F7F4EF", "ink": "#2D2A26", "green": "#2F7D5A",
    "sage": "#8BAA8B", "lemon": "#EEDC5B", "card": "#FFFFFF",
}

def apply_theme(page_title="Ferment Station", icon="🥤"):
    st.set_page_config(page_title=page_title, page_icon=icon, layout="wide")
    st.markdown(f"""
    <style>
      .block-container {{ max-width: 1400px; padding-top: 1rem; padding-bottom: 3rem; }}
      h1,h2,h3,h4,h5 {{ color:{COLORS['ink']}; letter-spacing:.2px; }}
      .section-title {{
        display:flex; align-items:center; gap:.5rem; padding:.4rem .8rem;
        background:{COLORS['sage']}22; border-left:6px solid {COLORS['sage']};
        border-radius:14px; margin:.2rem 0 1rem 0;
      }}
      .kpi {{
        background:{COLORS['card']}; border:1px solid #0001;
        border-left:6px solid {COLORS['green']}; border-radius:14px; padding:16px;
      }}
      .kpi .t {{ font-size:.9rem; color:#555; margin-bottom:6px; }}
      .kpi .v {{ font-size:1.5rem; font-weight:700; color:{COLORS['ink']}; }}
      div.stButton > button:first-child {{ background:{COLORS['green']}; color:#fff; border:none; border-radius:12px; }}
    </style>
    """, unsafe_allow_html=True)

def section(title: str, emoji=""):
    t = f"{emoji} {title}" if emoji else title
    st.markdown(f'<div class="section-title"><h2 style="margin:0">{t}</h2></div>', unsafe_allow_html=True)

def kpi(title: str, value: str):
    st.markdown(f'<div class="kpi"><div class="t">{title}</div><div class="v">{value}</div></div>', unsafe_allow_html=True)

# ---------- Images helpers ----------
IMG_EXTS = (".png", ".jpg", ".jpeg", ".webp", ".gif")

def slugify(s: str) -> str:
    s = str(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return s

IMG_EXTS = (".png", ".jpg", ".jpeg", ".webp", ".gif")

def find_image_path(images_dir: str, sku: str = None, flavor: str = None):
    """
    Ordre:
      0) assets/image_map.csv (canonical -> filename). Si filename sans extension, on essaie .jpg/.jpeg/.png/.webp/.gif
      1) Par SKU (CITR-33.ext puis CITR.ext)
      2) Par slug du goût (ex: mangue-passion.ext)
    """
    import os, csv, unicodedata, re as _re

    def _norm_key(s: str) -> str:
        s = str(s or "")
        s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
        s = _re.sub(r"\s+", " ", s).strip().lower()
        return s

    # 0) mapping CSV
    map_csv = os.path.join(images_dir, "image_map.csv")
    if os.path.exists(map_csv) and flavor:
        for sep in (",", ";"):
            try:
                d = {}
                with open(map_csv, "r", encoding="utf-8") as f:
                    rdr = csv.DictReader(f, delimiter=sep)
                    if not rdr.fieldnames:
                        continue
                    cols = {c.lower(): c for c in rdr.fieldnames}
                    if "canonical" in cols and "filename" in cols:
                        for row in rdr:
                            cano = (row.get(cols["canonical"]) or "").strip()
                            fn   = (row.get(cols["filename"])  or "").strip()
                            if cano and fn:
                                d[_norm_key(cano)] = fn
                        break
            except Exception:
                pass
        fn = d.get(_norm_key(flavor)) if 'd' in locals() else None
        if fn:
            p = os.path.join(images_dir, fn)
            if os.path.splitext(fn)[1] == "":  # pas d'extension
                for ext in IMG_EXTS:
                    p_try = p + ext
                    if os.path.exists(p_try):
                        return p_try
            if os.path.exists(p):
                return p

    # 1) SKU
    if sku:
        for ext in IMG_EXTS:
            p = os.path.join(images_dir, f"{sku}{ext}")
            if os.path.exists(p):
                return p
        base_root = _re.sub(r"-\d+$", "", sku)
        for ext in IMG_EXTS:
            p = os.path.join(images_dir, f"{base_root}{ext}")
            if os.path.exists(p):
                return p

    # 2) slug du goût
    if flavor:
        from .design import slugify  # si slugify est dans ce fichier, sinon adapte
        s = slugify(flavor)
        for ext in IMG_EXTS:
            p = os.path.join(images_dir, f"{s}{ext}")
            if os.path.exists(p):
                return p

    return None


import os, base64
from io import BytesIO
from PIL import Image

import os, base64
from io import BytesIO
from PIL import Image

def load_image_bytes(path: str):
    """
    Retourne :
    - bytes PNG (préféré)
    - ou data-URL base64 (fallback)
    """
    if not path or not os.path.exists(path):
        return None
    ext = os.path.splitext(path)[1].lower()
    try:
        im = Image.open(path).convert("RGBA")
        buf = BytesIO()
        im.save(buf, format="PNG")
        return buf.getvalue()
    except Exception:
        try:
            with open(path, "rb") as f:
                raw = f.read()
            mime = {
                ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                ".png": "image/png", ".webp": "image/webp", ".gif": "image/gif",
            }.get(ext, "image/octet-stream")
            b64 = base64.b64encode(raw).decode("ascii")
            return f"data:{mime};base64,{b64}"
        except Exception:
            return None

-Fichier = data.py :
import os, yaml, pandas as pd
from functools import lru_cache

CONFIG_DEFAULT = {
    "data_files": {
        "main_table": "data/production.xlsx",
        "flavor_map": "data/flavor_map.csv",
    },
    "images_dir": "assets",
}

def load_config() -> dict:
    path = "config.yaml"
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return {**CONFIG_DEFAULT, **(yaml.safe_load(f) or {})}
    return CONFIG_DEFAULT

@lru_cache(maxsize=1)
def get_paths():
    cfg = load_config()
    return (
        cfg["data_files"]["main_table"],
        cfg["data_files"]["flavor_map"],
        cfg["images_dir"],
    )

@lru_cache(maxsize=2)
def read_table():
    main_table, _, _ = get_paths()
    import os, pandas as pd

    if not os.path.exists(main_table):
        # Pas de fichier -> DataFrame vide
        return pd.DataFrame()

    lower = main_table.lower()
    try:
        if lower.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            # Formats Excel modernes -> openpyxl
            return pd.read_excel(main_table, engine="openpyxl", header=None)
        elif lower.endswith(".xls"):
            # Ancien Excel -> xlrd (nécessite xlrd dans requirements)
            return pd.read_excel(main_table, engine="xlrd", header=None)
        elif lower.endswith((".csv", ".txt")):
            # CSV/TXT du repo (séparateur ; si besoin adapte)
            try:
                return pd.read_csv(main_table, sep=";", engine="python", header=None)
            except Exception:
                return pd.read_csv(main_table, sep=",", engine="python", header=None)
        else:
            # Fallback: on tente openpyxl puis xlrd
            try:
                return pd.read_excel(main_table, engine="openpyxl", header=None)
            except Exception:
                return pd.read_excel(main_table, engine="xlrd", header=None)
    except Exception as e:
        # On remonte une table vide pour que l'accueil n’explose pas,
        # et on affiche l’erreur côté pages quand on relira le fichier proprement.
        return pd.DataFrame()

@lru_cache(maxsize=2)
def read_flavor_map():
    _, flavor_map, _ = get_paths()
    if not os.path.exists(flavor_map):
        return pd.DataFrame(columns=["name","canonical"])
    # essaie différents séparateurs si besoin
    try:
        return pd.read_csv(flavor_map, encoding="utf-8")
    except Exception:
        return pd.read_csv(flavor_map, encoding="utf-8", sep=";")

-DOSSIER 5 = assets 
(voir photo) 
- images qui s'affiche pas mais tkt 

- Fichier = image_map.csv : 
canonical,filename
Gingembre,GING.jpg
Mangue Passion,MAPA.jpg
Menthe citron vert,MECV.jpg
Original,ORIG.jpg
Pêche,PECH.jpg
Infusion Mélisse,MELI.jpg
Infusion menthe poivrée,INMP.jpg
