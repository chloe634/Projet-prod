# -*- coding: utf-8 -*-
# Ferment Station — Optimiseur + UI "Symbiose" (Streamlit)
# Requis: streamlit, pandas, numpy, pillow, openpyxl
# (facultatif) st-aggrid, plotly non nécessaires ici
#
# Lancer: streamlit run app.py

import io
import os
import re
from io import BytesIO
from pathlib import Path
from typing import Optional, List

import numpy as np
import pandas as pd
from PIL import Image
import streamlit as st

# =========================
# ---------- UI : palette & styles "Symbiose"
# =========================
COLORS = {
    "bg": "#F7F4EF",    # écru / papier
    "ink": "#2D2A26",   # gris charbon
    "green": "#2F7D5A", # vert forêt
    "sage": "#8BAA8B",  # vert sauge
    "lemon": "#EEDC5B", # accent "vitamines"
    "card": "#FFFFFF",
}

st.set_page_config(
    page_title="Optimiseur de production — 64 hL / 1–2 goûts",
    page_icon="🥤",
    layout="wide"
)

st.markdown(
    f"""
    <style>
      .block-container {{
        max-width: 1400px;
        padding-top: 1rem;
        padding-bottom: 3rem;
      }}
      body, .markdown-text-container, .stText, .stMarkdown p {{
        color: {COLORS['ink']};
      }}
      h1, h2, h3, h4, h5 {{
        color: {COLORS['ink']};
        letter-spacing: .2px;
      }}
      .section-title {{
        display:flex; align-items:center; gap:.5rem;
        padding:.4rem .8rem;
        background:{COLORS['sage']}22;
        border-left:6px solid {COLORS['sage']};
        border-radius:14px;
        margin:.2rem 0 1rem 0;
      }}
      .kpi {{
        background:{COLORS['card']};
        border:1px solid #00000010;
        border-left:6px solid {COLORS['green']};
        border-radius:14px;
        padding:16px;
      }}
      .kpi .t {{ font-size:.9rem; color:#555; margin-bottom:6px; }}
      .kpi .v {{ font-size:1.5rem; font-weight:700; color:{COLORS['ink']}; }}
      div.stButton > button:first-child {{
        background:{COLORS['green']}; color:#fff; border:none; border-radius:12px;
      }}
    </style>
    """,
    unsafe_allow_html=True
)

# =========================
# ---- Fix common encoding mishaps (Ã©, �, etc.) ----
# =========================
ACCENT_CHARS = "éèêëàâäîïôöùûüçÉÈÊËÀÂÄÎÏÔÖÙÛÜÇ"

# remplacements ciblés si le caractère d'origine est perdu (�)
CUSTOM_REPLACEMENTS = {
    "M�lisse": "Mélisse",
    "poivr�e": "poivrée",
    "P�che": "Pêche",
}

def _looks_better(a: str, b: str) -> bool:
    # decide if b has more accented chars than a
    def score(s): return sum(ch in ACCENT_CHARS for ch in s)
    return score(b) > score(a)

def fix_text(s) -> str:
    """Repair strings like 'MÃ©lisse' or 'poivr�e' -> 'Mélisse' / 'poivrée'."""
    if s is None:
        return ""
    if not isinstance(s, str):
        s = str(s)
    s0 = s

    # Case 1: UTF-8 bytes decoded as latin-1 → 'MÃ©lisse'
    try:
        s1 = s0.encode("latin1").decode("utf-8")
        if _looks_better(s0, s1):
            s0 = s1
    except Exception:
        pass

    # Case 2: exact known bad forms
    if s0 in CUSTOM_REPLACEMENTS:
        return CUSTOM_REPLACEMENTS[s0]

    # Case 3: replacement char '�' present → best-effort FR: replace by 'é'
    if "�" in s0:
        s0 = s0.replace("�", "é")

    return s0


# =========================
# Optimiseur de production — 64 hL / 1–2 goûts (Py 3.9)
# =========================

ALLOWED_FORMATS = {(12, 0.33), (6, 0.75), (4, 0.75)}
ROUND_TO_CARTON = True
VOL_TOL = 0.02
EPS = 1e-9
DEFAULT_WINDOW_DAYS = 60  # fallback si B2 introuvable

# ---------- Sidebar (Paramètres) ----------
with st.sidebar:
    st.header("Paramètres de production")
    volume_cible = st.number_input(
        "Volume cible (hL)",
        min_value=1.0, value=64.0, step=1.0,
        help="Si 1 goût: volume PAR goût. Si 2 goûts: volume TOTAL partagé."
    )
    nb_gouts = st.selectbox("Nombre de goûts simultanés", [1, 2], index=0)
    repartir_pro_rv = st.checkbox(
        "Répartition par formats au prorata des vitesses de vente",
        value=True,
        help="Sinon: répartition égale entre formats d'un même goût."
    )
    st.markdown("---")
    st.subheader("Prix moyen par hL")
    price_hL = st.number_input(
        "Prix moyen (€/hL)",
        min_value=0.0, value=500.0, step=10.0, format="%.0f",
        help="Prix moyen par hectolitre, indépendant des formats (utilisé pour le calcul des pertes)."
    )
    st.markdown("---")
    st.subheader("Navigation")
    page = st.radio(
        "Aller à",
        ["Tableau de production", "Optimisation & pertes", "Fiche de ramasse", "Paramètres"],
        index=0
    )

st.title("🥤 Optimiseur de production — 64 hL / 1–2 goûts")
st.caption("Fenêtre (jours) lue automatiquement en **B2**. Les calculs s’effectuent par **goût canonique** (mapping CSV).")

# ---------- Upload Excel ----------
uploaded = st.file_uploader("Dépose ton fichier Excel (.xlsx/.xls)", type=["xlsx", "xls"])
if uploaded is None:
    st.info("💡 Charge un fichier Excel pour commencer.")
    st.stop()

# ---------- Outils Excel / parsing ----------
def detect_header_row(df_raw: pd.DataFrame) -> int:
    must = {"Produit", "Stock", "Quantité vendue", "Volume vendu (hl)", "Quantité disponible", "Volume disponible (hl)"}
    for i in range(min(10, len(df_raw))):
        if must.issubset(set(str(x).strip() for x in df_raw.iloc[i].tolist())):
            return i
    return 0

def rows_to_keep_by_fill(excel_bytes: bytes, header_idx: int) -> List[bool]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    start_row = header_idx + 2
    keep: List[bool] = []
    for r in range(start_row, ws.max_row + 1):
        is_black = False
        for cell in ws[r]:
            fill = cell.fill
            if fill and fill.fill_type:
                rgb = getattr(getattr(fill, "fgColor", None), "rgb", None) or getattr(getattr(fill, "start_color", None), "rgb", None)
                if rgb and rgb[-6:].upper() == "000000":
                    is_black = True
                    break
        keep.append(not is_black)
    return keep

def parse_days_from_b2(value) -> Optional[int]:
    try:
        if isinstance(value, (int, float)) and not pd.isna(value):
            v = int(round(float(value)));  return v if v > 0 else None
        if value is None:
            return None
        s = str(value).strip()
        m = re.search(r"(\d+)\s*(?:j|jour|jours)\b", s, flags=re.IGNORECASE)
        if m:
            v = int(m.group(1));  return v if v > 0 else None
        date_pat = r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}).*?(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})"
        m2 = re.search(date_pat, s)
        if m2:
            d1 = pd.to_datetime(m2.group(1), dayfirst=True, errors="coerce")
            d2 = pd.to_datetime(m2.group(2), dayfirst=True, errors="coerce")
            if pd.notna(d1) and pd.notna(d2):
                days = int((d2 - d1).days)
                return days if days > 0 else None
        m3 = re.search(r"\b(\d{1,4})\b", s)
        if m3:
            v = int(m3.group(1));  return v if v > 0 else None
    except Exception:
        return None
    return None

def read_input_excel_and_period(uploaded_file):
    file_bytes = uploaded_file.read()
    raw = pd.read_excel(io.BytesIO(file_bytes), header=None)
    header_idx = detect_header_row(raw)
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_idx)
    keep_mask = rows_to_keep_by_fill(file_bytes, header_idx)
    if len(keep_mask) < len(df):
        keep_mask = keep_mask + [True] * (len(df) - len(keep_mask))
    df = df.iloc[[i for i, k in enumerate(keep_mask) if k]].reset_index(drop=True)
    # lecture de B2
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
        b2_val = ws["B2"].value
        wd = parse_days_from_b2(b2_val)
    except Exception:
        wd = None
    return df, (wd if wd and wd > 0 else DEFAULT_WINDOW_DAYS), file_bytes

# ---------- Mapping produits → goûts canoniques ----------
def load_flavor_map(uploaded_file=None) -> pd.DataFrame:
    """
    Charge flavor_map.csv en étant tolérant à l'encodage (utf-8, utf-8-sig, cp1252, latin1)
    et au séparateur (',' ';' tab '|'). Colonnes attendues : name, canonical.
    Nettoie aussi les accents via fix_text().
    """
    def _clean_cols(df: pd.DataFrame) -> pd.DataFrame:
        lower = {c.lower(): c for c in df.columns}
        keep = []
        if "name" in lower:
            keep.append(lower["name"])
        if "canonical" in lower:
            keep.append(lower["canonical"])
        if not keep:
            return pd.DataFrame(columns=["name", "canonical"])

        df = df[keep].copy()
        df.columns = ["name", "canonical"]

        # nettoyage valeurs + accents
        df = df.dropna(subset=["name", "canonical"])
        df["name"] = df["name"].astype(str).str.strip().map(fix_text)
        df["canonical"] = df["canonical"].astype(str).str.strip().map(fix_text)

        # enlève lignes vides après strip/fix
        df = df[(df["name"] != "") & (df["canonical"] != "")]
        return df

    # sources: upload (si présent) puis fichier du repo
    sources = []
    if uploaded_file is not None:
        sources.append(uploaded_file)
    p = Path(__file__).parent / "flavor_map.csv"
    if p.exists():
        sources.append(p)

    encodings = ["utf-8", "utf-8-sig", "cp1252", "latin1"]
    seps = [",", ";", "\t", "|"]

    for src in sources:
        for enc in encodings:
            for sep in seps:
                try:
                    fm = pd.read_csv(src, encoding=enc, sep=sep, engine="python")
                    fm = _clean_cols(fm)
                    if {"name", "canonical"}.issubset(fm.columns) and len(fm):
                        return fm
                except Exception:
                    continue
    # fallback: DF vide avec bonnes colonnes
    return pd.DataFrame(columns=["name", "canonical"])


def apply_canonical_flavor(df: pd.DataFrame, fm: pd.DataFrame) -> pd.DataFrame:
    """Mapping Produit -> Goût canonique (casse/espaces/accents normalisés + fallback flou)."""
    out = df.copy()

    # normalise & répare accents côté produits
    out["Produit"] = out["Produit"].astype(str).map(fix_text)
    out["Produit_norm"] = out["Produit"].str.strip()

    if len(fm):
        fm = fm.dropna(subset=["name", "canonical"]).copy()
        fm["name_norm"] = fm["name"].astype(str).map(fix_text).str.strip().str.lower()
        fm["canonical"] = fm["canonical"].astype(str).map(fix_text).str.strip()

        m_exact = dict(zip(fm["name_norm"], fm["canonical"]))
        keys = list(m_exact.keys())

        import difflib as _difflib
        def to_canonical(prod: str) -> str:
            s = str(prod).strip()
            k = s.lower()
            if k in m_exact:
                return m_exact[k]
            try:
                close = _difflib.get_close_matches(k, keys, n=1, cutoff=0.92)
                if close:
                    return m_exact[close[0]]
            except Exception:
                pass
            return s  # défaut: libellé produit normalisé

        out["GoutCanon"] = out["Produit_norm"].map(to_canonical)
    else:
        out["GoutCanon"] = out["Produit_norm"]

    # répare accents côté sortie également
    out["GoutCanon"] = out["GoutCanon"].astype(str).map(fix_text).str.strip()
    return out


# ---------- Parsing "Stock" et formats ----------
def parse_stock(text: str):
    if pd.isna(text): return np.nan, np.nan
    s = str(text)
    nb = None
    for pat in [
        r"(?:Carton|Caisse|Colis)\s+de\s*(\d+)",
        r"(\d+)\s*[x×]\s*Bouteilles?",
        r"(\d+)\s*[x×]",
        r"(\d+)\s+Bouteilles?",
    ]:
        m = re.search(pat, s, flags=re.IGNORECASE)
        if m:
            try: nb = int(m.group(1)); break
            except: pass
    vol_l = None
    m_l = re.findall(r"(\d+(?:[.,]\d+)?)\s*[lL]", s)
    if m_l: vol_l = float(m_l[-1].replace(",", "."))
    else:
        m_cl = re.findall(r"(\d+(?:[.,]\d+)?)\s*c[lL]", s)
        if m_cl: vol_l = float(m_cl[-1].replace(",", ".")) / 100.0
    if nb is None or vol_l is None:
        m_combo = re.search(r"(\d+)\s*[x×]\s*(\d+(?:[.,]\d+)?)+\s*([lc]l?)", s, flags=re.IGNORECASE)
        if m_combo:
            try:
                nb2 = int(m_combo.group(1))
                val = float(m_combo.group(2).replace(",", "."))
                unit = m_combo.group(3).lower()
                vol2 = val if unit.startswith("l") else val/100.0
                if nb is None: nb = nb2
                if vol_l is None: vol_l = vol2
            except: pass
    if (nb is None or np.isnan(nb)) and vol_l is not None and abs(vol_l - 0.75) <= VOL_TOL:
        if re.search(r"(?:\b4\s*[x×]\b|Carton\s+de\s*4\b|4\s+Bouteilles?)", s, flags=re.IGNORECASE):
            nb = 4
    return (float(nb) if nb is not None else np.nan,
            float(vol_l) if vol_l is not None else np.nan)

def safe_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")

def is_allowed_format(nb_bottles, vol_l, stock_txt: str) -> bool:
    if pd.isna(nb_bottles) or pd.isna(vol_l):
        if re.search(r"(?:\b4\s*[x×]\s*75\s*c?l\b|\b4\s+Bouteilles?\b.*75\s*c?l)", stock_txt, flags=re.IGNORECASE):
            nb_bottles = 4; vol_l = 0.75
        else:
            return False
    nb_bottles = int(nb_bottles); vol_l = float(vol_l)
    for nb_ok, vol_ok in ALLOWED_FORMATS:
        if nb_bottles == nb_ok and abs(vol_l - vol_ok) <= VOL_TOL:
            return True
    return False

# --------- Blocage de nan et "Autres (coffrets, goodies...)" ----------
BLOCKED_LABELS_EXACT = {"Autres (coffrets, goodies...)"}
BLOCKED_LABELS_LOWER = {"nan", "none", ""}

def sanitize_gouts(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["GoutCanon"] = out["GoutCanon"].astype(str).str.strip()
    mask = ~out["GoutCanon"].str.lower().isin(BLOCKED_LABELS_LOWER)
    mask &= ~out["GoutCanon"].isin(BLOCKED_LABELS_EXACT)
    return out.loc[mask].reset_index(drop=True)

# ---------- Cœur de calcul PRODUCTION ----------
def compute_plan(df_in, window_days, volume_cible, nb_gouts, repartir_pro_rv, manual_keep, exclude_list):
    required = ["Produit", "GoutCanon", "Stock", "Quantité vendue", "Volume vendu (hl)", "Quantité disponible", "Volume disponible (hl)"]
    miss = [c for c in required if c not in df_in.columns]
    if miss: raise ValueError(f"Colonnes manquantes: {miss}")

    df = df_in[required].copy()
    for c in ["Quantité vendue", "Volume vendu (hl)", "Quantité disponible", "Volume disponible (hl)"]:
        df[c] = safe_num(df[c])

    parsed = df["Stock"].apply(parse_stock)
    df[["Bouteilles/carton", "Volume bouteille (L)"]] = pd.DataFrame(parsed.tolist(), index=df.index)
    mask_allowed = df.apply(lambda r: is_allowed_format(r["Bouteilles/carton"], r["Volume bouteille (L)"], str(r["Stock"])), axis=1)
    df = df.loc[mask_allowed].reset_index(drop=True)

    df["Volume/carton (hL)"] = (df["Bouteilles/carton"] * df["Volume bouteille (L)"]) / 100.0
    df = df.dropna(subset=["GoutCanon", "Volume/carton (hL)", "Volume vendu (hl)", "Volume disponible (hl)"]).reset_index(drop=True)

    df_all_formats = df.copy()  # pour transparence & pertes (avant filtrages prod)

    # Exclusions production
    if exclude_list:
        ex = {s.strip() for s in exclude_list}
        df = df[~df["GoutCanon"].astype(str).str.strip().isin(ex)]

    # Sélection manuelle (production)
    if manual_keep:
        keep = {g.strip() for g in manual_keep}
        df = df[df["GoutCanon"].astype(str).str.strip().isin(keep)]

    # Sélection intelligente (autonomie + ventes)
    agg = df.groupby("GoutCanon").agg(
        ventes_hl=("Volume vendu (hl)", "sum"),
        stock_hl=("Volume disponible (hl)", "sum")
    )
    agg["vitesse_j"] = agg["ventes_hl"] / max(float(window_days), 1.0)
    agg["jours_autonomie"] = np.where(agg["vitesse_j"] > 0, agg["stock_hl"] / agg["vitesse_j"], np.inf)
    agg["score_urgence"] = agg["vitesse_j"] / (agg["jours_autonomie"] + EPS)
    agg = agg.sort_values(by=["score_urgence", "jours_autonomie", "ventes_hl"], ascending=[False, True, False])

    if not manual_keep:
        gouts_cibles = agg.index.tolist()[:nb_gouts]
    else:
        gouts_cibles = sorted(set(df["GoutCanon"]))
        if len(gouts_cibles) > nb_gouts:
            order = [g for g in agg.index if g in gouts_cibles]
            gouts_cibles = order[:nb_gouts]

    df_selected = df[df["GoutCanon"].isin(gouts_cibles)].copy()
    if len(gouts_cibles) == 0:
        raise ValueError("Aucun goût sélectionné (tout a peut-être été exclu).")

    # Calculs de production
    df_calc = df_selected.copy()
    if nb_gouts == 1:
        df_calc["Somme ventes (hL) par goût"] = df_calc.groupby("GoutCanon")["Volume vendu (hl)"].transform("sum")
        if repartir_pro_rv:
            df_calc["r_i"] = np.where(df_calc["Somme ventes (hL) par goût"] > 0,
                                      df_calc["Volume vendu (hl)"] / df_calc["Somme ventes (hL) par goût"], 0.0)
        else:
            df_calc["r_i"] = 1.0 / df_calc.groupby("GoutCanon")["GoutCanon"].transform("count")

        df_calc["G_i (hL)"] = df_calc["Volume disponible (hl)"]
        df_calc["G_total (hL) par goût"] = df_calc.groupby("GoutCanon")["G_i (hL)"].transform("sum")
        df_calc["Y_total (hL) par goût"] = df_calc["G_total (hL) par goût"] + float(volume_cible)
        df_calc["X_th (hL)"] = df_calc["r_i"] * df_calc["Y_total (hL) par goût"] - df_calc["G_i (hL)"]

        df_calc["X_adj (hL)"] = 0.0
        for gout, grp in df_calc.groupby("GoutCanon"):
            x = grp["X_th (hL)"].to_numpy(float)
            r = grp["r_i"].to_numpy(float)
            x = np.maximum(x, 0.0)
            deficit = float(volume_cible) - x.sum()
            if deficit > 1e-9:
                r = np.where(r > 0, r, 0); s = r.sum()
                x = x + (deficit * (r / s) if s > 0 else deficit / len(x))
            x = np.where(x < 1e-9, 0.0, x)
            df_calc.loc[grp.index, "X_adj (hL)"] = x
        cap_resume = f"{volume_cible:.2f} hL par goût"
    else:
        somme_ventes = df_calc["Volume vendu (hl)"].sum()
        if repartir_pro_rv and somme_ventes > 0:
            df_calc["r_i_global"] = df_calc["Volume vendu (hl)"] / somme_ventes
        else:
            df_calc["r_i_global"] = 1.0 / len(df_calc)

        df_calc["G_i (hL)"] = df_calc["Volume disponible (hl)"]
        G_total_all = df_calc["G_i (hL)"].sum()
        Y_total_all = G_total_all + float(volume_cible)
        df_calc["X_th (hL)"] = df_calc["r_i_global"] * Y_total_all - df_calc["G_i (hL)"]

        x = np.maximum(df_calc["X_th (hL)"].to_numpy(float), 0.0)
        deficit = float(volume_cible) - x.sum()
        if deficit > 1e-9:
            w = df_calc["r_i_global"].to_numpy(float); s = w.sum()
            x = x + (deficit * (w / s) if s > 0 else deficit / len(x))
        x = np.where(x < 1e-9, 0.0, x)
        df_calc["X_adj (hL)"] = x
        cap_resume = f"{volume_cible:.2f} hL au total (2 goûts)"

    # Cartons
    df_calc["Cartons à produire (exact)"] = df_calc["X_adj (hL)"] / df_calc["Volume/carton (hL)"]
    if ROUND_TO_CARTON:
        df_calc["Cartons à produire (arrondi)"] = np.floor(df_calc["Cartons à produire (exact)"] + 0.5).astype("Int64")
        df_calc["Volume produit arrondi (hL)"] = df_calc["Cartons à produire (arrondi)"] * df_calc["Volume/carton (hL)"]

    # Bouteilles (affichage: arrondi uniquement)
    df_calc["Bouteilles à produire (exact)"] = df_calc["Cartons à produire (exact)"] * df_calc["Bouteilles/carton"]
    if ROUND_TO_CARTON:
        df_calc["Bouteilles à produire (arrondi)"] = (
            df_calc["Cartons à produire (arrondi)"] * df_calc["Bouteilles/carton"]
        ).astype("Int64")

    # Sortie simplifiée (sans colonnes "exact")
    df_min = df_calc[[
        "GoutCanon", "Produit", "Stock",
        "Cartons à produire (arrondi)",
        "Bouteilles à produire (arrondi)",
        "Volume produit arrondi (hL)"
    ]].sort_values(["GoutCanon", "Produit", "Stock"]).reset_index(drop=True)

    # Transparence sélection
    agg_full = df.groupby("GoutCanon").agg(
        ventes_hl=("Volume vendu (hl)", "sum"),
        stock_hl=("Volume disponible (hl)", "sum")
    )
    agg_full["vitesse_j"] = agg_full["ventes_hl"] / max(float(window_days), 1.0)
    agg_full["jours_autonomie"] = np.where(agg_full["vitesse_j"] > 0, agg_full["stock_hl"] / agg_full["vitesse_j"], np.inf)
    agg_full["score_urgence"] = agg_full["vitesse_j"] / (agg_full["jours_autonomie"] + EPS)
    sel_gouts = sorted(set(df_calc["GoutCanon"]))
    synth_sel = agg_full.loc[sel_gouts][["ventes_hl", "stock_hl", "vitesse_j", "jours_autonomie", "score_urgence"]].copy()
    synth_sel = synth_sel.rename(columns={
        "ventes_hl": "Ventes 2 mois (hL)",
        "stock_hl": "Stock (hL)",
        "vitesse_j": "Vitesse (hL/j)",
        "jours_autonomie": "Autonomie (jours)",
        "score_urgence": "Score urgence"
    })

    return df_min, cap_resume, sel_gouts, synth_sel, df_calc, df_all_formats

# ---------- Pertes (projection 7 jours, indépendant des exclusions) ----------
def compute_losses_table_v48(
    df_in_all: pd.DataFrame,
    window_days: float,
    price_hL: float
) -> pd.DataFrame:
    """
    Pertes de CA par goût si NON produit pendant **7 jours**.
    Perte (€) = max(0, Demande_7j(hL) - Stock(hL)) × Prix_moyen(€/hL).
    Retourne toujours un DataFrame (éventuellement vide).
    """
    out_cols = ["Goût", "Demande 7 j (hL)", "Stock (hL)", "Manque sur 7 j (hL)", "Prix moyen (€/hL)", "Perte (€)"]
    if df_in_all is None or not isinstance(df_in_all, pd.DataFrame) or df_in_all.empty:
        return pd.DataFrame(columns=out_cols)

    df = df_in_all.copy()
    if "GoutCanon" not in df.columns:
        return pd.DataFrame(columns=out_cols)

    # numériques
    for c in ["Quantité vendue", "Volume vendu (hl)", "Quantité disponible", "Volume disponible (hl)"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    # filtre goûts indésirables
    df["GoutCanon"] = df["GoutCanon"].astype(str).str.strip()
    bad_lower = {"nan", "none", ""}
    df = df[~df["GoutCanon"].str.lower().isin(bad_lower)]
    df = df[df["GoutCanon"] != "Autres (coffrets, goodies...)"]
    if df.empty:
        return pd.DataFrame(columns=out_cols)

    jours = max(float(window_days), 1.0)
    agg = df.groupby("GoutCanon", as_index=False).agg(
        ventes_hL=("Volume vendu (hl)", "sum"),
        stock_hL=("Volume disponible (hl)", "sum"),
    )
    if agg.empty:
        return pd.DataFrame(columns=out_cols)

    agg["vitesse_hL_j"] = agg["ventes_hL"] / jours
    agg["Demande 7 j (hL)"] = 7.0 * agg["vitesse_hL_j"]
    agg["Stock (hL)"] = agg["stock_hL"]
    agg["Manque sur 7 j (hL)"] = np.clip(agg["Demande 7 j (hL)"] - agg["Stock (hL)"], a_min=0.0, a_max=None)
    agg["Prix moyen (€/hL)"] = float(price_hL)
    agg["Perte (€)"] = (agg["Manque sur 7 j (hL)"] * agg["Prix moyen (€/hL)"]).round(0)

    pertes = agg.rename(columns={"GoutCanon": "Goût"})[
        ["Goût", "Demande 7 j (hL)", "Stock (hL)", "Manque sur 7 j (hL)", "Prix moyen (€/hL)", "Perte (€)"]
    ]
    pertes["Goût"] = pertes["Goût"].map(fix_text)  # défensif : corrige accents en sortie
    pertes["Demande 7 j (hL)"] = pertes["Demande 7 j (hL)"].round(2)
    pertes["Stock (hL)"] = pertes["Stock (hL)"].round(2)
    pertes["Manque sur 7 j (hL)"] = pertes["Manque sur 7 j (hL)"].round(2)
    pertes["Prix moyen (€/hL)"] = pertes["Prix moyen (€/hL)"].round(0)

    return pertes.sort_values("Perte (€)", ascending=False).reset_index(drop=True)

# ---------- Lecture + période ----------
try:
    df_in_raw, window_days, file_bytes = read_input_excel_and_period(uploaded)
except Exception as e:
    st.error(f"Erreur de lecture : {e}")
    st.stop()
st.info(f"📅 Fenêtre détectée (B2) : **{window_days} jours** (défaut {DEFAULT_WINDOW_DAYS} si non détecté).")

# ---------- Mapping (depuis le repo uniquement) ----------
flavor_map_df = load_flavor_map()
df_in = apply_canonical_flavor(df_in_raw, flavor_map_df)

# --- normalize accents on both Produit & GoutCanon for display ---
df_in["Produit"] = df_in["Produit"].map(fix_text)
df_in["GoutCanon"] = df_in["GoutCanon"].map(fix_text)

df_in = sanitize_gouts(df_in)  # supprime 'nan' et 'Autres (coffrets, goodies...)'

# ---------- UI : exclusions / manuel ----------
with st.sidebar:
    raw_gouts = pd.Series(df_in.get("GoutCanon", pd.Series(dtype=str)))

    vals = (
        raw_gouts
        .dropna()
        .astype(str)
        .map(fix_text)     # corrige les 'MÃ©lisse', 'poivr�e', etc.
        .str.strip()
    )

    # blacklist pour supprimer 'nan', 'none', vide, et "Autres (coffrets, goodies...)"
    blacklist_lower = {"nan", "none", ""}
    vals = vals[~vals.str.lower().isin(blacklist_lower)]
    vals = vals[vals != "Autres (coffrets, goodies...)"]

    all_gouts = sorted(vals.unique().tolist())

    excluded_gouts = st.multiselect(
        "🚫 Exclure certains goûts (canoniques)", 
        options=all_gouts, 
        default=[]
    )

    use_manual = st.checkbox("Sélection manuelle du/des goût.s à produire", value=False)
    manual_keep = None
    if use_manual:
        manual_keep = st.multiselect(
            "Choisis les goûts à produire",
            options=[g for g in all_gouts if g not in excluded_gouts],
            default=[]
        )

# ---------- Calcul principal (production) ----------
try:
    df_min, cap_resume, gouts_cibles, synth_sel, df_selected_calc, df_all_formats = compute_plan(
        df_in=df_in,
        window_days=window_days,
        volume_cible=volume_cible,
        nb_gouts=nb_gouts,
        repartir_pro_rv=repartir_pro_rv,
        manual_keep=manual_keep,
        exclude_list=excluded_gouts
    )
except Exception as e:
    st.error(f"Erreur de calcul : {e}")
    st.stop()

# ---------- Pertes (projection 7 jours) ----------
pertes_tous_aucune_prod = compute_losses_table_v48(df_in, window_days, price_hL)

# =========================
# ---------- UI Helpers (images & KPIs)
# =========================
IMG_EXTS = (".png", ".jpg", ".jpeg", ".webp", ".gif")

def _slug(s:str)->str:
    s = fix_text(str(s)).lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s

def _maybe_sku_from_label(lbl:str):
    m = re.search(r"\b([A-Z]{3,5}-\d{2,3})\b", str(lbl))
    return m.group(1) if m else None

def _find_image_for(gout:str, base_dir="assets", sku:str=None):
    # 1) priorité fichier nommé par SKU (ex: assets/ORIG-33.png)
    if sku:
        for ext in IMG_EXTS:
            p = os.path.join(base_dir, f"{sku}{ext}")
            if os.path.exists(p): return p
    # 2) sinon fichier nommé par slug du goût (ex: assets/kefir-citron.png)
    key = _slug(gout)
    for ext in IMG_EXTS:
        p = os.path.join(base_dir, f"{key}{ext}")
        if os.path.exists(p): return p
    return None

def _img_bytes(path:str):
    if not path or not os.path.exists(path): return None
    im = Image.open(path).convert("RGBA")
    buf = BytesIO()
    im.save(buf, format="PNG")
    return buf.getvalue()

def kpi(title, value):
    st.markdown(f'<div class="kpi"><div class="t">{title}</div><div class="v">{value}</div></div>', unsafe_allow_html=True)


# =========================
# ---------- PAGES / AFFICHAGES
# =========================

# Prépare "df_prod" à partir du résultat de calcul (df_min)
df_prod = df_min.copy()

# Ajoute images (bytes) par ligne
df_prod["SKU_guess"] = df_prod["Produit"].apply(_maybe_sku_from_label)
df_prod["Image"] = [
    _img_bytes(_find_image_for(g, sku=_maybe_sku_from_label(p)))
    for g, p in zip(df_prod["GoutCanon"], df_prod["Produit"])
]

# Normalise numériques (évite NaN d'affichage)
if "Bouteilles à produire (arrondi)" in df_prod.columns:
    df_prod["Bouteilles à produire (arrondi)"] = df_prod["Bouteilles à produire (arrondi)"].fillna(0).astype("Int64")
if "Cartons à produire (arrondi)" in df_prod.columns:
    df_prod["Cartons à produire (arrondi)"] = df_prod["Cartons à produire (arrondi)"].fillna(0).astype("Int64")

# KPIs
try:
    total_btl = int(pd.to_numeric(df_prod.get("Bouteilles à produire (arrondi)"), errors="coerce").fillna(0).sum())
except Exception:
    total_btl = 0
total_vol = float(pd.to_numeric(df_prod.get("Volume produit arrondi (hL)"), errors="coerce").fillna(0).sum())
crit = 0
if isinstance(synth_sel, pd.DataFrame) and not synth_sel.empty and "Autonomie (jours)" in synth_sel.columns:
    crit = int((pd.to_numeric(synth_sel["Autonomie (jours)"], errors="coerce") <= 2).sum())

# === ROUTAGE DES PAGES ===
if page == "Tableau de production":
    st.markdown('<div class="section-title"><h2 style="margin:0">📦 Tableau de production</h2></div>', unsafe_allow_html=True)

    c1, c2, c3, c4 = st.columns(4)
    with c1: kpi("Total bouteilles à produire", f"{total_btl:,}".replace(",", " "))
    with c2: kpi("Volume total (hL)", f"{total_vol:.2f}")
    with c3: kpi("Goûts sélectionnés", f"{len(gouts_cibles)}")
    with c4: kpi("Produits critiques (≤2j)", f"{crit}")

    st.markdown("#### Vue détaillée (avec images)")
    st.dataframe(
        df_prod[[
            "Image", "GoutCanon", "Produit", "Stock",
            "Cartons à produire (arrondi)", "Bouteilles à produire (arrondi)", "Volume produit arrondi (hL)"
        ]],
        use_container_width=True,
        column_config={
            "Image": st.column_config.ImageColumn("Image", width="small"),
            "GoutCanon": "Goût",
            "Volume produit arrondi (hL)": st.column_config.NumberColumn(format="%.2f"),
        },
        hide_index=True
    )

    with st.expander("Pourquoi ces goûts ? (autonomie & ventes)"):
        st.dataframe(
            synth_sel.style.format({
                "Ventes 2 mois (hL)": "{:.2f}",
                "Stock (hL)": "{:.2f}",
                "Vitesse (hL/j)": "{:.3f}",
                "Autonomie (jours)": lambda v: '∞' if np.isinf(v) else f"{v:.1f}",
                "Score urgence": "{:.6f}",
            }),
            use_container_width=True
        )

elif page == "Optimisation & pertes":
    st.markdown('<div class="section-title"><h2 style="margin:0">📉 Optimisation & pertes</h2></div>', unsafe_allow_html=True)
    st.caption("Calcul pertes si on **ne produit rien** pendant 7 jours.")

    colA, colB = st.columns([2,1])
    with colA:
        if isinstance(pertes_tous_aucune_prod, pd.DataFrame) and not pertes_tous_aucune_prod.empty:
            st.dataframe(
                pertes_tous_aucune_prod[["Goût","Demande 7 j (hL)","Stock (hL)","Manque sur 7 j (hL)","Prix moyen (€/hL)","Perte (€)"]]
                .style.format({
                    "Demande 7 j (hL)": "{:.2f}",
                    "Stock (hL)": "{:.2f}",
                    "Manque sur 7 j (hL)": "{:.2f}",
                    "Prix moyen (€/hL)": "€{:,.0f}",
                    "Perte (€)": "€{:,.0f}",
                }),
                use_container_width=True, hide_index=True
            )
        else:
            st.info("Aucune perte estimée sur 7 jours (données insuffisantes ou stock suffisant).")
    with colB:
        perte_totale_aucune = float(pertes_tous_aucune_prod["Perte (€)"].sum()) if isinstance(pertes_tous_aucune_prod, pd.DataFrame) and not pertes_tous_aucune_prod.empty else 0.0
        kpi("Perte totale estimée (7 j)", f"€{perte_totale_aucune:,.0f}")

elif page == "Fiche de ramasse":
    st.markdown('<div class="section-title"><h2 style="margin:0">🚚 Fiche de ramasse</h2></div>', unsafe_allow_html=True)
    st.info("Espace réservé — dis-moi les colonnes à inclure (point de vente, tournée, SKU, quantité, conditionnement, notes) et je branche la génération PDF/Impression ici.")

    st.button("Générer la fiche (prochainement)")

elif page == "Paramètres":
    st.markdown('<div class="section-title"><h2 style="margin:0">⚙️ Paramètres</h2></div>', unsafe_allow_html=True)
    st.caption("Nom des images conseillés : par SKU (ex. ORIG-33.png) ou par goût (ex. kefir-citron.png). Place-les dans `assets/`.")
    st.write("Formats d’images reconnus :", ", ".join(IMG_EXTS))
    st.markdown("- **Astuce** : si aucune image n’est trouvée, la colonne s’affiche vide. Tu peux ajouter les visuels plus tard sans changer le code.")
